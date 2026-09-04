
import os
import openpyxl
import re
from datetime import datetime
from werkzeug.utils import secure_filename
import gc

def ordinal(n):
    """Convert number to ordinal (1→1st, 2→2nd, 3→3rd, etc.)"""
    try:
        n = int(n)
        if 11 <= (n % 100) <= 13:
            return f"{n}th"
        suffixes = {1: "st", 2: "nd", 3: "rd"}
        return f"{n}{suffixes.get(n % 10, 'th')}"
    except:
        return str(n)

def get_week_ranges(first_start, first_end, last_start, last_end):
    """Calculates week ranges dynamically based on the first and last week input."""
    try:
        f_s, f_e = int(float(first_start)), int(float(first_end))
        l_s, l_e = int(float(last_start)), int(float(last_end))
        
        weeks = [(f_s, f_e)]
        curr_start = f_e + 1
        while curr_start < l_s:
            curr_end = curr_start + 6
            if curr_end >= l_s:
                curr_end = l_s - 1
            if curr_start <= curr_end:
                weeks.append((curr_start, curr_end))
            curr_start = curr_end + 1
            
        weeks.append((l_s, l_e))
        flat_list = []
        for ws, we in weeks:
            flat_list.extend([ws, we])
        return weeks, flat_list
    except:
        return [], []

def safe_float(v):
    if v is None: return 0.0
    if isinstance(v, (int, float)): return float(v)
    try:
        clean_v = re.sub(r'[^\d.]', '', str(v))
        return float(clean_v) if clean_v else 0.0
    except:
        return 0.0

def process_zomato_pay(invoice_files, template_path, output_dir, update_progress=None, 
                       client_name="", month="", first_start=None, first_end=None, 
                       last_start=None, last_end=None, forced_filename=None):
    """
    Ultra-optimized Zomato Pay reconciliation with refined logic.
    """
    temp_files = []
    try:
        if update_progress: update_progress(5)

        # 1. Prepare Output Workbook
        if not os.path.exists(template_path):
            return None, f"Template file not found at {template_path}"
        
        out_wb = openpyxl.load_workbook(template_path)
        ws_calc = out_wb["Zpay Calculations"] if "Zpay Calculations" in out_wb.sheetnames else out_wb.create_sheet("Zpay Calculations")
        ws_ads = out_wb["Zpay Ads"] if "Zpay Ads" in out_wb.sheetnames else out_wb.create_sheet("Zpay Ads")
        
        # Clear existing content efficiently (avoid delete_rows which hog memory)
        for ws in [ws_calc, ws_ads]:
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
                for cell in row:
                    cell.value = None

        processed_data = [] # Stores Transaction Summary
        ads_data = [] # Stores Ad Summary

        # 2. Fast Input Reading (Read-Only)
        for idx, file in enumerate(invoice_files):
            filename = secure_filename(file.filename)
            temp_path = os.path.join(output_dir, f"temp_zpay_{filename}")
            file.save(temp_path)
            temp_files.append(temp_path)

            wb_in = openpyxl.load_workbook(temp_path, read_only=True, data_only=True)
            
            # Capture Transactions - search for header row dynamically
            if "Transactions summary" in wb_in.sheetnames:
                ws_in = wb_in["Transactions summary"]
                # Scanning rows 1-10 for headers
                header_row_idx = -1
                for r in range(1, 11):
                    row_vals = [str(c).lower() if c else "" for c in next(ws_in.iter_rows(min_row=r, max_row=r, values_only=True))]
                    if any("date" in h for h in row_vals) and any("bill" in h for h in row_vals):
                        header_row_idx = r
                        break
                
                if header_row_idx != -1:
                    # Include headers as the first row in processed_data for find_col
                    rows = list(ws_in.iter_rows(min_row=header_row_idx, values_only=True))
                    for row in rows:
                        if any(row): processed_data.append(row)
            
            # Capture Ads - search for header row dynamically
            if "Additions & deductions" in wb_in.sheetnames:
                ws_in = wb_in["Additions & deductions"]
                header_row_idx = -1
                for r in range(1, 6):
                    row_vals = [str(c).lower() if c else "" for c in next(ws_in.iter_rows(min_row=r, max_row=r, values_only=True))]
                    if any("type" in h for h in row_vals) and any("amount" in h for h in row_vals):
                        header_row_idx = r
                        break
                
                if header_row_idx != -1:
                    rows = list(ws_in.iter_rows(min_row=header_row_idx, values_only=True))
                    for row in rows:
                        if any(row): ads_data.append(row)
            wb_in.close()
        
        if update_progress: update_progress(35)

        # 3. Batch Write to Template (Fast Append)
        # Create 14 row gap as requested
        for _ in range(14): ws_calc.append([])
        for row in processed_data: ws_calc.append(row)
        
        for _ in range(5): ws_ads.append([])
        for row in ads_data: ws_ads.append(row)

        if update_progress: update_progress(60)

        # 4. Calculation Mapping (Headers are the first row of collected data)
        headers = [str(h).strip().lower() if h else "" for h in processed_data[0]] if processed_data else []
        def find_col(possible_names):
            for name in possible_names:
                for idx, h in enumerate(headers):
                    if name.lower() in h: return idx
            return -1

        col_date = find_col(["date and time", "date", "time", "transaction date"])
        col_bill = find_col(["bill amount", "order amount", "total bill"])
        col_discount = find_col(["instant discount", "discount amount", "total discount", "promo share"])
        col_promo = find_col(["promo share", "restaurant share"])
        col_comm = find_col(["commission amount", "commission", "zomato commission"])
        col_tip = find_col(["tips", "tip amount"])
        col_net = find_col(["net receivable", "payout", "settlement amount"])

        if col_date == -1 or col_bill == -1:
            return None, "Required date or bill columns missing in Transactions summary."

        # Strict Month Filtering
        month_map = {"january": 1, "february": 2, "march": 3, "april": 4, "may": 5, "june": 6,
                     "july": 7, "august": 8, "september": 9, "october": 10, "november": 11, "december": 12}
        target_month_num = month_map.get(month.lower())

        weeks, _ = get_week_ranges(first_start, first_end, last_start, last_end)
        weekly_stats = {i: {'bill':0, 'disc':0, 'comm':0, 'tip':0, 'net':0} for i in range(len(weeks))}
        
        # Adjustments for prev/next month
        adj_prev_month = 0.0
        adj_next_month = 0.0

        # Performance: Loop over processed_data list directly (skip first row which is headers)
        for idx, row in enumerate(processed_data[1:]):
            date_val = row[col_date]
            if not date_val: continue
            
            day, m_num = None, None
            if isinstance(date_val, datetime):
                day, m_num = date_val.day, date_val.month
            else:
                parts = re.findall(r'\d+', str(date_val))
                if len(parts) >= 3:
                    if len(parts[0]) == 4: # YYYY-MM-DD
                        day, m_num = int(parts[2]), int(parts[1])
                    else: # DD-MM-YYYY
                        day, m_num = int(parts[0]), int(parts[1])

            if day is None: continue
            
            # Handle Adjustments (Prev/Next month) - Only if month differs
            if target_month_num and m_num != target_month_num:
                # Determine if previous or next month (with year boundary support)
                is_prev = False
                if target_month_num == 1 and m_num == 12: is_prev = True
                elif target_month_num == 12 and m_num == 1: is_prev = False
                elif m_num < target_month_num: is_prev = True
                else: is_prev = False

                if is_prev:
                    adj_prev_month += safe_float(row[col_net])
                else:
                    adj_next_month += safe_float(row[col_net])
                
                # Skip weekly distribution for adjustment rows
                continue

            for i, (ws, we) in enumerate(weeks):
                if ws <= day <= we:
                    stats = weekly_stats[i]
                    stats['bill'] += safe_float(row[col_bill])
                    # Fixed Logic: Discounts use direct sum to match yellow cell
                    stats['disc'] += (safe_float(row[col_discount]) if col_discount != -1 else 0) + \
                                     (safe_float(row[col_promo]) if col_promo != -1 else 0)
                    stats['comm'] += safe_float(row[col_comm])
                    stats['tip'] += safe_float(row[col_tip]) if col_tip != -1 else 0
                    stats['net'] += safe_float(row[col_net])
                    
                    # Mark week for debugger (idx already +1 from [1:] but ws_calc row starts at 16)
                    ws_calc.cell(row=16+idx, column=len(row)+1).value = f"W{i+1}"
                    break

        # 5. Inject Weekly Results into Row 2-6 (G onwards)
        calc_results = {i: stats for i, stats in weekly_stats.items()}
        for i in range(len(weeks)):
            stats = weekly_stats[i]
            x_col = 7 + i # G=7, H=8, etc.
            ws_calc.cell(row=1, column=x_col).value = f"Week {i+1} Recon"
            ws_calc.cell(row=2, column=x_col).value = stats['bill'] * (100.0/105.0)
            ws_calc.cell(row=3, column=x_col).value = stats['disc'] * (100.0/105.0) # Apply 100/105 to Discounts
            ws_calc.cell(row=4, column=x_col).value = stats['comm'] * 1.18
            ws_calc.cell(row=5, column=x_col).value = stats['tip']
            ws_calc.cell(row=6, column=x_col).value = stats['net']

        # 6. Zpay Ads Logic
        ads_headers = [str(h).strip().lower() if h else "" for h in ads_data[0]] if ads_data else []
        col_ads_date = -1
        col_ads_amt = -1
        col_ads_type = -1
        for idx, h in enumerate(ads_headers):
            if "date" in h: col_ads_date = idx
            if "amount" in h: col_ads_amt = idx
            if "type" in h: col_ads_type = idx
        
        ads_weekly = {i: 0.0 for i in range(len(weeks))}
        ads_prev_month = 0.0
        ads_next_month = 0.0

        if col_ads_date != -1 and col_ads_amt != -1:
            # Skip first row which is headers
            for idx, row in enumerate(ads_data[1:]):
                date_val = row[col_ads_date]
                if not date_val: continue
                day, m_num = None, None
                if isinstance(date_val, datetime):
                    day, m_num = date_val.day, date_val.month
                else:
                    parts = re.findall(r'\d+', str(date_val))
                    if len(parts) >= 3:
                        if len(parts[0]) == 4: day, m_num = int(parts[2]), int(parts[1])
                        else: day, m_num = int(parts[0]), int(parts[1])
                
                if day is None: continue

                val = safe_float(row[col_ads_amt])
                # If it's a deduction (common for ads), it should be negative. 
                # If the file already shows it as negative, safe_float handles it.
                
                # Handle Adjustments for Ads - Only if month differs
                if target_month_num and m_num != target_month_num:
                    is_prev = False
                    if target_month_num == 1 and m_num == 12: is_prev = True
                    elif target_month_num == 12 and m_num == 1: is_prev = False
                    elif m_num < target_month_num: is_prev = True
                    else: is_prev = False

                    if is_prev:
                        ads_prev_month += val
                    else:
                        ads_next_month += val
                    continue

                for i, (ws_range, we_range) in enumerate(weeks):
                    if ws_range <= day <= we_range:
                        ads_weekly[i] += val
                        ws_ads.cell(row=7+idx, column=len(row)+1).value = f"W{i+1}"
                        break
        
        for i in range(len(weeks)):
            ws_ads.cell(row=1, column=7+i).value = f"W{i+1}"
            ws_ads.cell(row=2, column=7+i).value = ads_weekly[i]

        # 7. Final Mapping to Zomato Pay (Consolidated)
        if "Zomato Pay" in out_wb.sheetnames:
            ws_final = out_wb["Zomato Pay"]
            
            # Client name in B1, Month replacement in A2 and B2
            ws_final["B1"].value = client_name
            for cell_id in ["A2", "B2"]:
                current_val = str(ws_final[cell_id].value or "")
                if "month" in current_val.lower():
                    ws_final[cell_id].value = current_val.replace("Month", month).replace("month", month)
            
            # Paste Week Ranges starting at D5 (as requested)
            for i, (ws, we) in enumerate(weeks):
                label = f"{ordinal(ws)} to {ordinal(we)}"
                ws_final.cell(row=5, column=4+i).value = label

            last_week_col = 4 + len(weeks) - 1 # Dynamically find last week column

            # Determine last week column index (D=4, E=5, F=6, G=7, H=8)
            num_weeks = len(weeks)
            last_col_idx = 4 + num_weeks - 1

            for r in range(1, ws_final.max_row + 1):
                raw_cell_val = str(ws_final.cell(row=r, column=3).value or "")
                label = raw_cell_val.strip().lower()
                
                # Precise Adjustment Matching (Avoid possessives/plurals issues)
                clean_label = label.replace("'s", "").replace("  ", " ")
                
                if "opening week" in clean_label and "adjustment" in clean_label:
                    # Opening stock only in first week (column D = 4)
                    ws_final.cell(row=r, column=4).value = adj_prev_month + ads_prev_month
                    # Clear other week columns for this row
                    for c_off in range(1, num_weeks):
                        ws_final.cell(row=r, column=4+c_off).value = 0.0
                    continue
                
                elif "closing week" in clean_label and "adjustment" in clean_label:
                    # Closing stock only in the last week column
                    ws_final.cell(row=r, column=last_col_idx).value = adj_next_month + ads_next_month
                    # Clear previous week columns for this row
                    for c_off in range(0, num_weeks - 1):
                        ws_final.cell(row=r, column=4+c_off).value = 0.0
                    continue
                
                # Mapping Logic for other rows
                if "sales (exclusive of gst) before failed" in label:
                    for i in range(num_weeks):
                        ws_final.cell(row=r, column=4+i).value = calc_results[i]['bill'] * (100.0/105.0)
                
                elif "less: discounts" in label:
                    for i in range(num_weeks):
                        ws_final.cell(row=r, column=4+i).value = -(calc_results[i]['disc'] * (100.0/105.0))
                
                elif "add : tips" in label:
                    for i in range(num_weeks):
                        ws_final.cell(row=r, column=4+i).value = calc_results[i]['tip']
                
                elif "commission (inclusive of gst)" in label:
                    for i in range(num_weeks):
                        ws_final.cell(row=r, column=4+i).value = calc_results[i]['comm'] * 1.18
                
                elif "zomatopay ads" in label:
                    for i in range(num_weeks):
                        # Map ads to their respective weeks
                        ws_final.cell(row=r, column=4+i).value = -ads_weekly[i]

        # Save and Cleanup
        output_filename = forced_filename if forced_filename else f"Zomato_Pay_Recon_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        full_path = os.path.join(output_dir, output_filename)
        out_wb.save(full_path)
        out_wb.close()
        
        for f in temp_files:
            try: os.remove(f)
            except: pass
        
        if update_progress: update_progress(100)
        gc.collect()
        return output_filename, None

    except Exception as e:
        import traceback; traceback.print_exc()
        return None, str(e)
