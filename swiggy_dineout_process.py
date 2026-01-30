
import os
import openpyxl
import re
from datetime import datetime
import pandas as pd
from werkzeug.utils import secure_filename
import gc

def parse_date_range(date_str):
    """
    Parses a date range string like "01 October - 05 October" or "28 Sep - 05 Oct".
    Returns a tuple (start_date_obj, end_date_obj, original_string).
    """
    try:
        if not isinstance(date_str, str):
            return None, None, str(date_str)

        parts = date_str.split('-')
        if len(parts) != 2:
            return None, None, date_str
        
        start_str = parts[0].strip()
        end_str = parts[1].strip()
        
        formats = ['%d %B', '%d %b']
        
        start_date = None
        end_date = None
        
        for fmt in formats:
            try:
                start_date = datetime.strptime(start_str, fmt)
                start_date = start_date.replace(year=2024) 
                break
            except ValueError:
                continue

        for fmt in formats:
            try:
                end_date = datetime.strptime(end_str, fmt)
                end_date = end_date.replace(year=2024)
                break
            except ValueError:
                continue

        if start_date and end_date:
            if end_date < start_date:
                end_date = end_date.replace(year=start_date.year + 1)
                
        if not start_date or not end_date:
            return None, None, date_str
            
        return start_date, end_date, date_str
    except Exception as e:
        print(f"Error parsing date: {e}")
        return None, None, date_str

def extract_data_block_fast(ws):
    """
    Extracts data block using iterator for speed (read-only friendly).
    """
    start_row = None
    end_row = None
    rows_data = [] 
    
    current_row_idx = 0
    # Optimization: Only iterate columns B to F
    for row in ws.iter_rows(min_col=2, max_col=6, values_only=True): 
        current_row_idx += 1
        val_b = row[0] 
        
        if not start_row:
            if val_b and isinstance(val_b, str) and "Payout Invoice" in val_b:
                start_row = current_row_idx
                rows_data.append(row[:4]) # Capture Headers
                continue
                
        if start_row and not end_row:
            rows_data.append(row[:4]) 
            if val_b and isinstance(val_b, str) and "Net Payout" in val_b:
                 if "D" in val_b and "E" in val_b:
                    end_row = current_row_idx
                    break
    
    if start_row and end_row:
        return rows_data
    return None

def apply_sd_calculations(ws):
    """
    Performs the calculations in Column E based on labels in Column A.
    Also sums positive ad values in F1.
    Returns a dict of values for consolidation.
    """
    total_ads = 0.0
    extracted = {
        'order_total': 0.0,
        'discount': 0.0,
        'service_fee': 0.0,
        'ads': 0.0,
        'tip': 0.0
    }
    
    # Iterate through all rows in the sheet
    for row in ws.iter_rows(min_col=1, max_col=5, values_only=False):
        cell_a = row[0]
        cell_b = row[1]
        cell_c = row[2] 
        cell_d = row[3]
        cell_e = row[4] # Column E (Target)

        if not cell_a.value:
            continue

        label = str(cell_a.value).strip()
        val_d = 0.0
        try:
            val_d = float(cell_d.value) if cell_d.value is not None else 0.0
        except ValueError: 
            val_d = 0.0
            
        val_b = 0.0
        try:
            val_b = float(cell_b.value) if cell_b.value is not None else 0.0
        except ValueError:
            val_b = 0.0

        # Logic
        if "Order Total" in label:
            # Col E = Col D * 100/105
            res = val_d * (100.0 / 105.0)
            cell_e.value = res
            extracted['order_total'] = res
            
        elif "Total merchant discount" in label:
            # Col E = Col B * 100/105
            res = val_b * (100.0 / 105.0)
            cell_e.value = res
            extracted['discount'] = res
            
        elif label == "Tip": # Strict Match
            cell_e.value = val_d
            extracted['tip'] = val_d
            
        elif "Swiggy Platform Service Fee" in label:
            # Col E = Col D * 1.18 (GST)
            res = val_d * 1.18
            cell_e.value = res
            extracted['service_fee'] = res
            
        elif any(x in label.upper() for x in ["TOP CAROUSEL", "TOP_CAROUSEL", "LISTING AD", "LISTING_AD", "AD CAMPAIGN", "AD_CAMPAIGN"]):
            # Col E = -Col D (Turn to positive)
            positive_val = -val_d
            cell_e.value = positive_val
            total_ads += positive_val

    # Write Total Ads to F1
    ws['F1'] = total_ads
    extracted['ads'] = total_ads
    
    return extracted

def get_ordinal(n):
    """
    Returns the ordinal representation of a number (1st, 2nd, etc.)
    """
    if 11 <= (n % 100) <= 13:
        suffix = 'th'
    else:
        suffix = {1: 'st', 2: 'nd', 3: 'rd'}.get(n % 10, 'th')
    return str(n) + suffix

def format_range_ordinal(start_date, end_date):
    """
    Converts (start_date, end_date) to "1st to 5th" format.
    """
    if not start_date or not end_date:
        return ""
    return f"{get_ordinal(start_date.day)} to {get_ordinal(end_date.day)}"

def consolidate_swiggy_dineout(out_wb, sd_data_map, client_name="", month_name=""):
    """
    Maps values from SD subsheets to the 'Swiggy Dineout' consolidated sheet.
    """
    sheet_name = "Swiggy Dineout" 
    if sheet_name not in out_wb.sheetnames:
         for n in out_wb.sheetnames:
             if "swiggy" in n.lower() and "dineout" in n.lower():
                 sheet_name = n
                 break
    
    if sheet_name not in out_wb.sheetnames:
        return

    ws_con = out_wb[sheet_name]
    
    # --- Final Touches: Global Replacements ---
    if client_name:
        ws_con['A1'] = client_name
        
    # Month Replacement in A2
    current_a2 = str(ws_con['A2'].value or "")
    if "month" in current_a2.lower():
        ws_con['A2'].value = current_a2.replace("Month", month_name).replace("month", month_name)

    if month_name:
        # Search and replace "November" in the whole sheet (common placeholder)
        for row in ws_con.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and "November" in cell.value:
                    cell.value = cell.value.replace("November", month_name)

    # 1. Identify Target Rows
    row_mapping = {
        "Sales (Exclusive of GST)": {'key': 'order_total', 'col': 2},
        "less: Discounts": {'key': 'discount', 'col': 2, 'negate': True},
        "Swiggy Platform Service Fee": {'key': 'service_fee', 'col': 2},
        "add: Tips": {'key': 'tip', 'col': 2},
        "Carousel, High Priority, Banner": {'key': 'ads', 'col': 2} 
    }
    
    found_rows = {}
    for r_idx, row in enumerate(ws_con.iter_rows(min_row=1, max_row=100, min_col=1, max_col=2, values_only=True), 1):
        val_a = str(row[0]).strip() if row[0] else ""
        val_b = str(row[1]).strip() if row[1] else ""
        for label, config in row_mapping.items():
            check_val = val_b if config['col'] == 2 else val_a
            if label.lower() in check_val.lower():
                if config['key'] not in found_rows:
                    found_rows[config['key']] = {'row': r_idx, 'negate': config.get('negate', False)}
            
            if config['key'] == 'ads' and 'carousel' in check_val.lower() and 'banner' in check_val.lower():
                 if config['key'] not in found_rows:
                    found_rows[config['key']] = {'row': r_idx, 'negate': config.get('negate', False)}

    # 2. Map Data & Date Headers
    for sd_key, info in sd_data_map.items():
        try:
            sd_num = int(re.search(r'\d+', sd_key).group())
            target_col = 3 + (sd_num - 1)
            
            # --- Final Touches: Date Header at C5, D5, etc. ---
            formatted_date = format_range_ordinal(info['start'], info['end'])
            ws_con.cell(row=5, column=target_col).value = formatted_date
            
            data = info['extracted']
            for key, val in data.items():
                if key in found_rows:
                    row_info = found_rows[key]
                    final_val = val
                    if row_info['negate']:
                        final_val = -abs(val)
                    ws_con.cell(row=row_info['row'], column=target_col).value = final_val
                    
        except Exception as e:
            print(f"⚠️ Error mapping {sd_key}: {e}")

def process_swiggy_dineout(invoice_files, template_path, output_dir, update_progress=None, client_name="", month="", forced_filename=None):
    """
    Final optimized processing logic.
    """
    temp_files = []
    processed_items = []
    
    try:
        if update_progress: update_progress(10)

        for idx, file in enumerate(invoice_files):
            filename = secure_filename(file.filename)
            temp_path = os.path.join(output_dir, f"temp_{filename}")
            file.save(temp_path)
            temp_files.append(temp_path)
            
            try:
                wb = openpyxl.load_workbook(temp_path, read_only=True, data_only=True)
                start_date, end_date = datetime.max, datetime.max
                
                if 'Summary' in wb.sheetnames:
                    ws_sum = wb['Summary']
                    date_val = ws_sum['B18'].value 
                    if date_val:
                        s, e, _ = parse_date_range(date_val)
                        if s: start_date, end_date = s, e
                
                target_sheet = None
                for sname in wb.sheetnames:
                    if "payout" in sname.lower() and "invoice" in sname.lower():
                        target_sheet = wb[sname]
                        break
                if not target_sheet: target_sheet = wb.active
                
                data_content = extract_data_block_fast(target_sheet)
                processed_items.append({
                    'start': start_date,
                    'end': end_date,
                    'data': data_content,
                    'filename': filename
                })
                wb.close()
                del wb
            except Exception as e:
                print(f"❌ Error processing {filename}: {e}")
            
            if update_progress: 
                p = 10 + int((idx+1)/len(invoice_files)*40)
                update_progress(p)

        processed_items.sort(key=lambda x: x['start'])
        if update_progress: update_progress(60)

        if not os.path.exists(template_path):
             out_wb = openpyxl.Workbook()
        else:
            out_wb = openpyxl.load_workbook(template_path)
            
        consolidation_map = {} 
        current_sd_index = 1
        previous_end_date = None
        
        for idx, item in enumerate(processed_items):
            start, end, data, fname = item['start'], item['end'], item['data'], item['filename']
            
            if previous_end_date and start != datetime.max:
                gap_days = (start - previous_end_date).days
                if gap_days > 3:
                     gap_adjusted = max(0, gap_days)
                     missing = round(gap_adjusted / 3.5)
                     missing = max(int(missing), 0)
                     if missing > 0:
                        current_sd_index += missing
            
            sheet_name = f"SD{current_sd_index}"
            if data:
                if sheet_name in out_wb.sheetnames:
                    ws_out = out_wb[sheet_name]
                else:
                    ws_out = out_wb.create_sheet(sheet_name)
                    
                for r_i, row in enumerate(data):
                    for c_i, val in enumerate(row):
                        ws_out.cell(row=r_i+1, column=c_i+1).value = val
                        
                results = apply_sd_calculations(ws_out)
                consolidation_map[sheet_name] = {
                    'extracted': results,
                    'start': start,
                    'end': end
                }

            previous_end_date = end
            current_sd_index += 1
            if update_progress: 
                p = 60 + int((idx+1)/len(processed_items)*35)
                update_progress(p)

        consolidate_swiggy_dineout(out_wb, consolidation_map, client_name=client_name, month_name=month)

        output_filename = forced_filename if forced_filename else f"Swiggy_Dineout_Recon_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        full_path = os.path.join(output_dir, output_filename)
        out_wb.save(full_path)
        out_wb.close()
        
        for f in temp_files:
            try: os.remove(f)
            except: pass
        
        gc.collect()
        return output_filename, None
    except Exception as e:
        import traceback
        tb = traceback.format_exc()
        print(f"❌ Critical Error in Swiggy Dineout: {tb}")
        return None, f"Processing Error: {str(e)}"


