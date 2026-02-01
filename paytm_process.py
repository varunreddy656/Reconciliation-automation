import openpyxl
import pandas as pd
import shutil
import os
import gc
import re
from datetime import datetime
from process_invoices import calculate_week_structure, ordinal, parse

def get_safe_dimensions(sheet):
    """Safe way to get max_row and max_column in read_only mode"""
    max_r = sheet.max_row
    max_c = sheet.max_column
    if max_r is None or max_c is None:
        if max_c is None:
            for row in sheet.iter_rows(min_row=1, max_row=10):
                if len(row) > (max_c or 0):
                    max_c = len(row)
        if max_r is None:
            max_r = 0
            for row in sheet.iter_rows(values_only=True):
                max_r += 1
    return max_r or 0, max_c or 0

def safe_float(val):
    if val is None: return 0.0
    try:
        if isinstance(val, str):
            val = val.replace('₹', '').replace(',', '').replace("'", "").strip()
            if not val or val == '-': return 0.0
        return float(val)
    except:
        return 0.0

def process_paytm(
    invoice_path,
    template_path,
    output_path,
    client_name="Client",
    month="October",
    first_week_start=None,
    first_week_end=None,
    last_week_start=None,
    last_week_end=None,
    progress_callback=None
):
    """
    Paytm Reconciliation Logic.
    """
    try:
        if progress_callback: progress_callback(10)

        # 1. Setup Template
        shutil.copy2(template_path, output_path)
        recon = openpyxl.load_workbook(output_path)
        
        if "Paytm Calculations" not in recon.sheetnames:
            return {'success': False, 'message': 'Sheet "Paytm Calculations" not found in template'}
        if "Paytm Reconciliation" not in recon.sheetnames:
            return {'success': False, 'message': 'Sheet "Paytm Reconciliation" not found in template'}
            
        ws_calc = recon["Paytm Calculations"]
        ws_recon = recon["Paytm Reconciliation"]

        if progress_callback: progress_callback(20)

        # 2. Load Source Data
        if invoice_path.lower().endswith('.csv'):
            df_src = pd.read_csv(invoice_path)
        else:
            df_src = pd.read_excel(invoice_path)

        # 3. Paste data to Paytm Calculations starting A10
        # Headers at row 10
        headers = df_src.columns.tolist()
        for idx, h in enumerate(headers, 1):
            ws_calc.cell(row=10, column=idx).value = h
            
        # Data from row 11 onwards
        for r_idx, row in enumerate(df_src.values, 11):
            for c_idx, val in enumerate(row, 1):
                ws_calc.cell(row=r_idx, column=c_idx).value = val

        if progress_callback: progress_callback(40)

        # 4. Identify Columns in Row 10
        status_col, amount_col, commission_col, date_col = -1, -1, -1, -1
        for idx, h in enumerate(headers):
            h_clean = str(h).strip().lower()
            if h_clean == "status": status_col = idx
            if h_clean == "amount": amount_col = idx
            if h_clean == "commission": commission_col = idx
            if h_clean == "transaction_date": date_col = idx

        if -1 in [status_col, amount_col, commission_col, date_col]:
            return {'success': False, 'message': f'Required columns not found. Found: Status={status_col}, Amount={amount_col}, Commission={commission_col}, Date={date_col}'}

        # 5. Calculate Week Structure
        week_structure = calculate_week_structure(month, first_week_start, first_week_end, last_week_start, last_week_end)
        
        # 6. Aggregate Week Wise
        # G2, H2, etc for Amount (100/105)
        # G3, H3, etc for Commission (1.18)
        
        weekly_stats = {w['week_num']: {'amt': 0.0, 'comm': 0.0, 'label': w['label']} for w in week_structure}
        
        for idx, row in df_src.iterrows():
            # Check Status
            status = str(row.iloc[status_col]).strip().upper().replace("'", "")
            if status != "SUCCESS": continue
            
            # Check Date
            date_val = str(row.iloc[date_col]).replace("'", "").strip()
            try:
                row_date_dt = parse(date_val)
                row_date = row_date_dt.date()
            except:
                continue
                
            # Match to Week
            for week in week_structure:
                if week['start_date'].date() <= row_date <= week['end_date'].date():
                    wn = week['week_num']
                    weekly_stats[wn]['amt'] += safe_float(row.iloc[amount_col])
                    weekly_stats[wn]['comm'] += safe_float(row.iloc[commission_col])
                    break

        # 7. Write results to Paytm Calculations
        for wn, stats in weekly_stats.items():
            col = 7 + (wn - 1) # G, H, I...
            ws_calc.cell(row=2, column=col).value = stats['amt'] * 100.0 / 105.0
            ws_calc.cell(row=3, column=col).value = stats['comm'] * 1.18
            ws_calc.cell(row=1, column=col).value = f"Week {wn}"

        if progress_callback: progress_callback(70)

        # 8. Map to Paytm Reconciliation
        # Client Name in A1
        ws_recon["A1"].value = client_name

        # Month Replacement in A2
        current_a2 = str(ws_recon["A2"].value or "")
        if "month" in current_a2.lower():
            ws_recon["A2"].value = current_a2.replace("Month", month).replace("month", month)
        else:
            ws_recon["A2"].value = client_name
        
        # Week ranges starting C5
        for wn, stats in weekly_stats.items():
            col = 3 + (wn - 1) # C, D, E...
            ws_recon.cell(row=5, column=col).value = stats['label']
            
        # Sales (exclusive of GST) and Commission Mapping
        # Sales row (Amt): Row where col B contains "Sales (exclusive of GST)"
        # Commission row: Row where col B contains "Commission (Inclusive of GST)"
        amt_row, comm_row = -1, -1
        max_recon_r, _ = get_safe_dimensions(ws_recon)
        for r in range(1, max_recon_r + 1):
            lbl = str(ws_recon.cell(row=r, column=2).value or "").lower()
            if "sales (exclusive of gst)" in lbl and "failed" in lbl:
                amt_row = r
            if "commission (inclusive of gst)" in lbl:
                comm_row = r
                
        if amt_row != -1:
            for wn in weekly_stats:
                col = 3 + (wn - 1)
                source_cell = ws_calc.cell(row=2, column=7 + (wn - 1)).coordinate
                ws_recon.cell(row=amt_row, column=col).value = f"='Paytm Calculations'!{source_cell}"
                
        if comm_row != -1:
            for wn in weekly_stats:
                col = 3 + (wn - 1)
                source_cell = ws_calc.cell(row=3, column=7 + (wn - 1)).coordinate
                ws_recon.cell(row=comm_row, column=col).value = f"='Paytm Calculations'!{source_cell}"

        if progress_callback: progress_callback(90)

        recon.save(output_path)
        recon.close()
        gc.collect()

        if progress_callback: progress_callback(100)
        return {'success': True}

    except Exception as e:
        import traceback
        traceback.print_exc()
        return {'success': False, 'message': str(e)}
