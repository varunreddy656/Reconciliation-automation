import openpyxl
from pathlib import Path
import re
import shutil
import uuid
import os
import time

_original_print = print
def print(*args, **kwargs):
    try:
        _original_print(*args, **kwargs)
    except UnicodeEncodeError:
        clean_args = [
            str(a).encode('ascii', 'ignore').decode('ascii') if isinstance(a, str) else a
            for a in args
        ]
        _original_print(*clean_args, **kwargs)


# ----------------- Helper Functions -----------------

def extract_swiggy_start_day(filepath):
    filename = Path(filepath).name
    print(f"--- Extracting start day for {filename} ---", flush=True)
    
    # 1. Filename regex matching space or underscore: e.g. "01 Aug 2026", "10_Aug_2026", "01-08-2026"
    date_match = re.search(r'(\d{1,2})[\s_]+([A-Za-z]{3}|\d{1,2})[\s_]+\d{2,4}', filename)
    if date_match:
        extracted_day = int(date_match.group(1))
        print(f"✅ Found start day {extracted_day} from filename.", flush=True)
        return extracted_day

    # 2. Try simple digit extraction at start of filename e.g. "01_Swiggy..."
    num_match = re.search(r'(\d{1,2})', filename)
    
    # 3. Robust scan across Summary or first sheet
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
        sheet_names = wb.sheetnames
        sheet = wb["Summary"] if "Summary" in sheet_names else wb.active
        text_found = None
        
        for r in range(1, 25):
            for c in range(1, 10):
                val = str(sheet.cell(row=r, column=c).value or "").strip()
                if re.search(r"(\d{1,2})\s*(?:st|nd|rd|th)?\s*[a-zA-Z]{0,9}.*?[-to]+\s*(\d{1,2})", val, re.IGNORECASE):
                    text_found = val
                    print(f"🔍 Found date pattern at R{r}C{c}: '{val}'", flush=True)
                    break
            if text_found: break
                
        wb.close()
        if text_found:
            m = re.search(r"(\d{1,2})\s*.*?[-to]+\s*(\d{1,2})", text_found, re.IGNORECASE)
            if m:
                extracted_day = int(m.group(1))
                print(f"✅ Extracted start day {extracted_day} from text '{text_found}'.", flush=True)
                return extracted_day
    except Exception as e:
        print(f"⚠️ Error reading Excel file {filename}: {e}", flush=True)

    if num_match:
        extracted_day = int(num_match.group(1))
        if 1 <= extracted_day <= 31:
            print(f"✅ Fallback start day {extracted_day} from filename digits.", flush=True)
            return extracted_day

    print(f"⚠️ Fallback to start day 1 for {filename}.", flush=True)
    return 1


def detect_platform(fp):
    filename = Path(fp).name.lower()
    if "swiggy" in filename:
        return "Swiggy"
    try:
        wb = openpyxl.load_workbook(fp, data_only=True, read_only=True)
        sheets = [s.lower() for s in wb.sheetnames]
        wb.close()
        if any("order" in s for s in sheets) and any("summary" in s for s in sheets):
            return "Swiggy"
        if any("charge" in s or "deduction" in s for s in sheets):
            return "Swiggy"
    except Exception as e:
        print(f"❌ detect_platform failed to open {filename}: {e}", flush=True)
        return "Swiggy"
        
    return "Swiggy"


def clear_all_D_sheets(wb):
    sheets_to_remove = [sh for sh in wb.sheetnames if sh.startswith("D1W") or sh.startswith("D2W")]
    for sh_name in sheets_to_remove:
        std = wb[sh_name]
        wb.remove(std)
    print(f"Cleared {len(sheets_to_remove)} old D1W/D2W sheets.")


def ensure_sheet(wb, name):
    if name in wb.sheetnames:
        return wb[name]
    else:
        return wb.create_sheet(name)


def replace_month_in_sheets(wb, user_month):
    sheets_to_update = ['Summary', 'Cashflow', 'Profit statement', 'Discrepancies']
    for sheet_name in sheets_to_update:
        if sheet_name not in wb.sheetnames:
            print(f"Sheet '{sheet_name}' not found. Skipping.")
            continue
        sheet = wb[sheet_name]
        if sheet_name == 'Summary':
            cell_value = str(sheet['B2'].value) if sheet['B2'].value else ''
            if 'July' in cell_value:
                sheet['B2'].value = cell_value.replace('July', user_month)
                print(f"Summary sheet B2 updated: 'July' → '{user_month}'")
        else:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and 'July' in cell.value:
                        cell.value = cell.value.replace('July', user_month)
            print(f"Sheet '{sheet_name}': All 'July' replaced with '{user_month}'")


def ordinal(n):
    if 11 <= (n % 100) <= 13:
        return f"{n}th"
    suffixes = {1: "st", 2: "nd", 3: "rd"}
    return f"{n}{suffixes.get(n % 10, 'th')}"


def format_week_label(start, end):
    return f"{ordinal(start)} to {ordinal(end)}"


def generate_week_ranges(first_start, first_end, last_start, last_end, max_day=31):
    week_ranges = []
    week_ranges.append((first_start, first_end))
    current_start = first_end + 1
    # Fixed: use `current_start < last_start` (not `current_start + 6 < last_start`)
    # so partial middle weeks (e.g. 22-24) are not skipped
    while current_start < last_start:
        current_end = min(current_start + 6, last_start - 1)
        if current_start <= current_end:
            week_ranges.append((current_start, current_end))
        current_start = current_end + 1
    week_ranges.append((last_start, last_end))
    return week_ranges


def fast_clear_sheet(ws):
    if ws and ws.max_row > 0:
        try:
            ws.delete_rows(1, ws.max_row)
        except Exception:
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
                for cell in row:
                    cell.value = None

def copy_data(src, tgt, start_row):
    fast_clear_sheet(tgt)
    for r_idx, row in enumerate(src.iter_rows(min_row=start_row, values_only=True), 1):
        for c_idx, val in enumerate(row, 1):
            tgt.cell(row=r_idx, column=c_idx).value = val


def extract_total_orders(fp):
    try:
        wb = openpyxl.load_workbook(fp, data_only=True, read_only=True)
        sheet = wb["Summary"]
        total_orders_value = None
        # Robust search for Total Orders label in Column B
        for r in range(1, 100):
            val_b = str(sheet.cell(row=r, column=2).value or "").strip().lower()
            if "total orders (delivered + cancelled)" in val_b:
                total_orders_value = sheet.cell(row=r, column=3).value
                break
        wb.close()
        return total_orders_value
    except Exception as e:
        print(f"Error extracting Total Orders from {fp}: {e}")
        return None


def count_non_zero_complaints(sheet):
    """
    Count non-zero values in the Customer Complaints column (row 6 onwards).
    Returns the count as an integer.
    """
    complaints_col = None
    for col in range(1, sheet.max_column + 1):
        cell_val = sheet.cell(row=5, column=col).value
        if cell_val and "customer complaints" in str(cell_val).strip().lower():
            complaints_col = col
            break
    if complaints_col is None:
        print("Customer Complaints column not found in D1W sheet.")
        return 0
    count = 0
    for row_vals in sheet.iter_rows(min_row=6, values_only=True):
        if row_vals and len(row_vals) >= complaints_col:
            value = row_vals[complaints_col - 1]
            if isinstance(value, (int, float)) and not isinstance(value, bool) and value != 0:
                count += 1
    return count


def map_values_to_cashflow(wb, data1_sheet, week):
    cashflow = wb["Cashflow"]
    week_col = 3 + (week - 1)

    data2_sheet_name = f"D2W{week}"
    data2_sheet = wb[data2_sheet_name] if data2_sheet_name in wb.sheetnames else None

    mapping = {
        "Item sales (Delivered orders)": (["Item Total"], 2, "single"),
        "Add:- Packing charges": (["Packaging Charges"], 2, "single"),
        "Add:- Compensation paid for cancelled orders": (
            ["Total Customer Paid", "Complaint & Cancellation Charges"], 1, "sub"),
        "Less:- Discount": (["Restaurant Discounts", "Swiggy One Exclusive Offer Discount"], 2, "sum"),
        "Add:- GST 5%": (["GST Collected"], 2, "single"),
        "Swiggy One Fees": (["Swiggy One Fees"], 3, "single"),
        "Call Center Service Fees": (["Call Center Charges"], 3, "single"),
        "PocketHero Fee": (["Pocket Hero Fees"], 3, "single"),
        "Platform Fee": (["Commission"], 3, "single"),
        "Long Distance Fee": (["Long Distance Charges"], 3, "single"),
        "Merchant Cancellation Charges": (["Restaurant Cancellation Charges"], 3, "single"),
        "Paid by Restaurant": (["Customer Complaints"], 4, "single"),
        "TDS deduction for aggrigators": (["TDS"], 4, "single"),
        "TCS": (["TCS"], 4, "single"),
        "GST collected and paid by swiggy": (["GST Deduction"], 4, "single"),
        "Collection Charges": (["Payment Collection Charges"], 3, "single")
    }

    partial_match_keywords = {
        "Total Customer Paid": "Total Customer Paid",
        "Complaint & Cancellation Charges": ["Complaint", "Cancellation"],
        "Restaurant Discounts": "Restaurant Discount",
        "Swiggy One Exclusive Offer Discount": "Swiggy One",
        "TCS": "TCS"
    }

    headers = {str(data1_sheet.cell(row=5, column=c).value).strip(): c
               for c in range(1, data1_sheet.max_column + 1)
               if data1_sheet.cell(row=5, column=c).value}

    def find_column(header_name):
        h_clean = header_name.strip()
        if h_clean in headers:
            return headers[h_clean]
        if h_clean in partial_match_keywords:
            keyword = partial_match_keywords[h_clean]
            if isinstance(keyword, list):
                for actual_header, col in headers.items():
                    if all(kw.lower() in actual_header.lower() for kw in keyword):
                        return col
            else:
                for actual_header, col in headers.items():
                    if keyword.lower() in actual_header.lower():
                        return col
        return None

    for row in range(1, cashflow.max_row + 1):
        label = cashflow.cell(row=row, column=2).value
        if not label:
            continue
        label = str(label).strip()
        if label not in mapping:
            continue

        data_headers, data_row, operation = mapping[label]
        data_cells = []

        for h in data_headers:
            col = find_column(h)
            if col:
                data_cells.append(data1_sheet.cell(row=data_row, column=col))
            else:
                print(f"Warning: Header '{h.strip()}' not found in Data1 sheet for '{label}'")

        if not data_cells:
            print(f"Skipping '{label}' because no Data1 headers found")
            continue

        formula = ""
        if operation == "single":
            formula = f"='{data1_sheet.title}'!{data_cells[0].coordinate}"
        elif operation == "sum":
            formula = "=" + "+".join([f"'{data1_sheet.title}'!{c.coordinate}" for c in data_cells])
        elif operation == "sub":
            if len(data_cells) != 2:
                print(f"Skipping subtraction for '{label}' because requires exactly 2 cells")
                continue
            formula = f"='{data1_sheet.title}'!{data_cells[0].coordinate}-'{data1_sheet.title}'!{data_cells[1].coordinate}"

        cashflow.cell(row=row, column=week_col).value = formula

    if data2_sheet:
        for row in range(1, cashflow.max_row + 1):
            label = cashflow.cell(row=row, column=2).value
            if not label:
                continue
            label = str(label).strip()

            if label == "High Priority":
                # Logic: Sum multiple Ad types from Column C
                # Flexible matching: search for headers anywhere in Column A (case-insensitive)
                ad_headers = ["cost per click - ads", "search ads (cba)", "top picks - ads", "ads offers", "brandverse social media marketing", "hyper boost", "hyperboost", "unlimited ads"]
                found_cells = []
                for r in range(1, data2_sheet.max_row + 1):
                    val_a = str(data2_sheet.cell(row=r, column=1).value or "").strip().lower()
                    if any(h in val_a for h in ad_headers):
                        # Find value in Column B onwards until found
                        target_val_cell = None
                        for col_idx in range(2, data2_sheet.max_column + 1):
                            cell = data2_sheet.cell(row=r, column=col_idx)
                            val = cell.value
                            
                            # Check if the value is valid (numeric and not a date)
                            is_valid = False
                            if val is not None and val != "":
                                if isinstance(val, (int, float)) and not isinstance(val, bool):
                                    is_valid = True
                                elif isinstance(val, str):
                                    cleaned = val.strip()
                                    cleaned_num = cleaned.replace('₹', '').replace(',', '').strip()
                                    if cleaned_num:
                                        try:
                                            float(cleaned_num)
                                            is_valid = True
                                        except ValueError:
                                            pass
                            
                            if is_valid:
                                target_val_cell = cell
                                break
                        
                        if target_val_cell:
                            found_cells.append(f"'{data2_sheet.title}'!{target_val_cell.coordinate}")
                
                if found_cells:
                    formula = "=-(" + "+".join(found_cells) + ")"
                    target_cell = cashflow.cell(row=row, column=week_col)
                    target_cell.value = formula
                    target_cell.number_format = '0.00'
                    print(f"High Priority mapped as sum of {len(found_cells)} ad rows")
                else:
                    cashflow.cell(row=row, column=week_col).value = 0
                    print(f"Warning: No specific Ads found in {data2_sheet.title}")

            elif label == "Previous week Adjustments":
                # Logic: Find all 'Previous week outstanding' rows in Column A
                found_cells = []
                for r in range(1, data2_sheet.max_row + 1):
                    val_a = str(data2_sheet.cell(row=r, column=1).value or "").strip().lower()
                    if "previous week outstanding" in val_a:
                        # Find value in Column B onwards until found
                        for col_idx in range(2, data2_sheet.max_column + 1):
                            cell = data2_sheet.cell(row=r, column=col_idx)
                            if cell.value is not None and cell.value != "":
                                found_cells.append(f"'{data2_sheet.title}'!{cell.coordinate}")
                                break
                
                if found_cells:
                    # Use negative sign to make it positive (since Swiggy lists it as negative)
                    formula = "=-(" + "+".join(found_cells) + ")"
                    target_cell = cashflow.cell(row=row, column=week_col)
                    target_cell.value = formula
                    target_cell.number_format = '0.00' # Ensure it's a plain number without currency symbols
                    print(f"Previous week Adjustments mapped as sum of {len(found_cells)} rows from {data2_sheet.title}")
                else:
                    cashflow.cell(row=row, column=week_col).value = 0
                    print(f"Warning: 'Previous week outstanding' not found in {data2_sheet.title}")

            elif label == "Paid by Restaurant":
                found_cells = []
                for r in range(1, data2_sheet.max_row + 1):
                    val_a = str(data2_sheet.cell(row=r, column=1).value or "").strip().lower()
                    if "discount deduction - cashback" in val_a:
                        for col_idx in range(2, data2_sheet.max_column + 1):
                            cell = data2_sheet.cell(row=r, column=col_idx)
                            val = cell.value
                            if val is not None and val != "":
                                num_val = None
                                if isinstance(val, (int, float)):
                                    num_val = float(val)
                                elif isinstance(val, str):
                                    cleaned = val.replace(',', '').strip()
                                    cleaned_num = ''.join(c for c in cleaned if c.isdigit() or c == '.' or c == '-')
                                    if cleaned_num:
                                        try:
                                            num_val = float(cleaned_num)
                                        except ValueError:
                                            pass
                                
                                if num_val is not None:
                                    m_cell = data2_sheet.cell(row=r, column=13)
                                    m_cell.value = abs(num_val)
                                    m_cell.number_format = '0.00'
                                    found_cells.append(f"'{data2_sheet.title}'!{m_cell.coordinate}")
                                    break
                
                if found_cells:
                    target_cell = cashflow.cell(row=row, column=week_col)
                    existing_formula = str(target_cell.value or "")
                    addition = "".join([f"+{c}" for c in found_cells])
                    
                    if existing_formula.startswith("="):
                        target_cell.value = existing_formula + addition
                    elif existing_formula == "0" or not existing_formula:
                        target_cell.value = "=" + addition.lstrip("+")
                    else:
                        target_cell.value = "=" + str(target_cell.value) + addition
                    
                    target_cell.number_format = '0.00'
                    print(f"Added {len(found_cells)} 'Discount Deduction - Cashback' rows to Paid by Restaurant via col M")
    else:
        print(f"Warning: Data2 sheet '{data2_sheet_name}' not found for week {week}")

    print(f"Cashflow mapped for week {week}")


def perform_calculations_on_data1(wb, data1_sheet, week, recon_path):
    data1_sheet.insert_rows(1, 4)
    item_total_col = None
    order_status_col = None
    for col_num, cell in enumerate(data1_sheet[5], 1):
        if cell.value == 'Item Total': item_total_col = col_num
        if cell.value == 'Order Status': order_status_col = col_num
    if not item_total_col or not order_status_col:
        print("Required columns missing")
        return

    num_cols = max(1, data1_sheet.max_column - item_total_col + 1)
    delivered = [0] * num_cols
    cancelled = [0] * num_cols

    for row_vals in data1_sheet.iter_rows(min_row=6, values_only=True):
        if not row_vals or len(row_vals) < order_status_col:
            continue
        status = str(row_vals[order_status_col - 1] or "").strip().lower()
        if status not in ["delivered", "cancelled"]:
            continue
        target = delivered if status == "delivered" else cancelled
        for i in range(num_cols):
            col_idx = item_total_col - 1 + i
            if col_idx < len(row_vals):
                val = row_vals[col_idx]
                if isinstance(val, (int, float)) and not isinstance(val, bool):
                    target[i] += val

    for i, col in enumerate(range(item_total_col, data1_sheet.max_column + 1)):
        data1_sheet.cell(row=4, column=col).value = delivered[i] + cancelled[i]
        data1_sheet.cell(row=2, column=col).value = delivered[i]
        data1_sheet.cell(row=1, column=col).value = cancelled[i]

    for col in range(item_total_col, data1_sheet.max_column + 1):
        val = data1_sheet.cell(row=4, column=col).value
        data1_sheet.cell(row=3, column=col).value = val * 1.18 if isinstance(val, (int, float)) else 0

    print("Row1/Row2/Row3/Row4 calculations done for", data1_sheet.title)
    map_values_to_cashflow(wb, data1_sheet, week)


def copy_bank_sheet_to_recon(bank_wb, recon_wb, bank_sheet_name=None, recon_bank_name='BANK'):
    if recon_bank_name in recon_wb.sheetnames:
        std = recon_wb[recon_bank_name]
        recon_wb.remove(std)
    bank_sheet_name = bank_sheet_name or bank_wb.sheetnames[0]
    bank_sheet = bank_wb[bank_sheet_name]
    recon_bank = recon_wb.create_sheet(recon_bank_name)
    for i, row in enumerate(bank_sheet.iter_rows(values_only=True), 1):
        for j, val in enumerate(row, 1):
            recon_bank.cell(row=i, column=j).value = val


def extract_expected_receipt(fp):
    try:
        wb = openpyxl.load_workbook(fp, data_only=True, read_only=True)
        sheet = wb["Summary"]
        val = sheet["C14"].value
        wb.close()
        if isinstance(val, str):
            val = val.replace(",", "").replace("₹", "").strip()
            try:
                val = float(val)
            except Exception:
                return None
        if isinstance(val, (int, float)):
            return float(val)
        return None
    except Exception as e:
        print(f"Error extracting expected receipts from {fp}: {e}")
        return None


def map_bank_to_actual_receipts_from_invoice_summary(recon_wb, week_expected_map, tolerance=10):
    if "BANK" not in recon_wb.sheetnames or "Cashflow" not in recon_wb.sheetnames:
        print("BANK or Cashflow sheet missing; skipping bank mapping.")
        return
    bank_sheet = recon_wb["BANK"]
    cashflow = recon_wb["Cashflow"]
    deposit_col = None
    deposit_labels = ["deposit amt.", "credit amount(inr)", "credit", "deposit", "amount"]
    for col in range(1, bank_sheet.max_column + 1):
        header = str(bank_sheet.cell(row=1, column=col).value).lower().strip() if bank_sheet.cell(row=1,
                                                                                                  column=col).value else ""
        for label in deposit_labels:
            if label.lower() in header:
                deposit_col = col
                print(f"Found deposit column at column {col} with header: '{header}'")
                break
        if deposit_col:
            break
    if deposit_col is None:
        print("Could not find deposit column.")
        return
    bank_deposits = []
    for row in range(2, bank_sheet.max_row + 1):
        val = bank_sheet.cell(row=row, column=deposit_col).value
        orig_val = val
        if isinstance(val, str):
            cleaned = val.replace(",", "").replace("\u200c", "").replace("₹", "").replace(" ", "")
            cleaned = ''.join(c for c in cleaned if c.isdigit() or c == '.' or c == '-')
            try:
                val = float(cleaned)
            except Exception:
                val = None
        if isinstance(val, (int, float)) and val != 0:
            bank_deposits.append((val, row))
            print(f"Bank deposit found at row {row} (raw: {orig_val}): {val}")
    actual_row = None
    for row in range(1, cashflow.max_row + 1):
        cell_val = str(cashflow.cell(row=row, column=2).value).strip().lower() if cashflow.cell(row=row,
                                                                                                column=2).value else ""
        if "actual receipts" in cell_val:
            actual_row = row
            break
    if not actual_row:
        print("Actual Receipts row not found.")
        return
    for week, expected in week_expected_map.items():
        col = 3 + (week - 1)
        print(f"Week {week}: expected receipt from invoice summary = {expected}")
        closest, min_diff = None, float('inf')
        for deposit, _ in bank_deposits:
            diff = abs(deposit - expected)
            if diff <= tolerance and diff < min_diff:
                closest, min_diff = deposit, diff
        if closest is not None:
            cashflow.cell(row=actual_row, column=col).value = closest
            print(f"Week {week}: mapped deposit {closest} (diff={min_diff}) at Cashflow row {actual_row} col {col}")
        else:
            # Assign 0 to clear out unmapped weeks
            cashflow.cell(row=actual_row, column=col).value = 0
            print(f"Week {week}: no matching bank deposit found within ±{tolerance}, setting cell to 0")

def safe_delete_bank_file(bank_file_path, retries=10, wait=0.5):
    bank_file_abs = os.path.abspath(bank_file_path)
    for attempt in range(retries):
        try:
            os.remove(bank_file_abs)
            print(f"Deleted bank file: {bank_file_abs}")
            return True
        except PermissionError as e:
            print(f"Attempt {attempt + 1}/{retries}: PermissionError while deleting bank file. Retrying...")
            time.sleep(wait)
        except Exception as e:
            print(f"Attempt {attempt + 1}/{retries}: Error while deleting bank file: {e}")
            break
    print(f"Could not delete bank file: {bank_file_abs}")
    return False


def convert_summary_row_to_numbers(summary_sheet, row=6, start_col='C', end_col='G'):
    """
    Convert cells in Summary sheet row 6 (C6:G6) from text to numbers.
    """
    col_map = {'C': 3, 'D': 4, 'E': 5, 'F': 6, 'G': 7}
    start_col_idx = col_map.get(start_col, 3)
    end_col_idx = col_map.get(end_col, 7)

    for col in range(start_col_idx, end_col_idx + 1):
        cell = summary_sheet.cell(row=row, column=col)
        if cell.value is not None:
            try:
                # Convert to numeric
                if isinstance(cell.value, str):
                    cleaned = cell.value.replace(',', '').replace(' ', '').strip()
                    numeric_val = float(cleaned)
                    if numeric_val.is_integer():
                        numeric_val = int(numeric_val)
                else:
                    numeric_val = float(cell.value)
                    if numeric_val.is_integer():
                        numeric_val = int(numeric_val)

                cell.value = numeric_val
                cell.number_format = '0'  # Integer format
                print(f"Converted C{row} to numeric: {numeric_val}")
            except Exception as e:
                print(f"Could not convert cell at col {col} row {row}: {e}")


def add_notepoints_based_on_bank(recon_wb, week_ranges, bank_file_path):
    cashflow = recon_wb["Cashflow"]
    discrepancies = recon_wb["Discrepancies"] if "Discrepancies" in recon_wb.sheetnames else None

    # Clear prior notes at B38 (Cashflow) and B24 (Discrepancies)
    cashflow.cell(row=38, column=2).value = None
    if discrepancies:
        discrepancies.cell(row=24, column=2).value = None

    actual_row = None
    for row in range(1, cashflow.max_row + 1):
        cell_val = str(cashflow.cell(row=row, column=2).value).strip().lower() if cashflow.cell(row=row, column=2).value else ""
        if "actual receipts" in cell_val:
            actual_row = row
            break
    if not actual_row:
        print("Actual Receipts row not found, skipping note point addition.")
        return

    mapped_weeks = set()
    for week_idx in range(1, len(week_ranges) + 1):
        col = 3 + (week_idx - 1)
        cell_val = cashflow.cell(row=actual_row, column=col).value
        # Consider non-zero, non-empty numeric values as mapped
        if cell_val not in [None, "", 0]:
            mapped_weeks.add(week_idx)

    all_weeks = set(range(1, len(week_ranges) + 1))
    unmapped_weeks = sorted(all_weeks - mapped_weeks)

    print(f"Debug: Weeks total: {all_weeks}, Mapped: {mapped_weeks}, Unmapped: {unmapped_weeks}")

    def ordinal(n):
        if 11 <= (n % 100) <= 13:
            return f"{n}th"
        suffixes = {1: "st", 2: "nd", 3: "rd"}
        return f"{n}{suffixes.get(n % 10, 'th')}"

    # Compose notepoint
    if not bank_file_path:
        note = "1. Due to absence of Bank, actual receipts could not be mapped."
    elif unmapped_weeks:
        week_names = [f"{ordinal(w)} week" for w in unmapped_weeks]
        if len(week_names) == 1:
            weeks_str = week_names[0]
        else:
            weeks_str = ", ".join(week_names[:-1]) + " and " + week_names[-1]
        note = f"1. Bank credits of {weeks_str} will be credited in the next month."
    else:
        print("All weeks mapped; no notepoint needed.")
        return

    print(f"Adding note: {note}")

    cashflow.cell(row=38, column=2).value = note
    if discrepancies:
        discrepancies.cell(row=24, column=2).value = note


def copy_images_from_template(template_path, output_path):
    """
    Copy images and drawings from template to output workbook.
    """
    from openpyxl.drawing.image import Image as XLImage
    from copy import copy

    template_wb = openpyxl.load_workbook(template_path)
    output_wb = openpyxl.load_workbook(output_path)

    for sheet_name in template_wb.sheetnames:
        if sheet_name not in output_wb.sheetnames:
            continue

        template_sheet = template_wb[sheet_name]
        output_sheet = output_wb[sheet_name]

        # Copy images/drawings
        if hasattr(template_sheet, '_images'):
            for img in template_sheet._images:
                output_sheet.add_image(img)

        if hasattr(template_sheet, '_drawing'):
            output_sheet._drawing = template_sheet._drawing

    output_wb.save(output_path)
    template_wb.close()


def process_invoices_web(
        invoice_folder_path,
        template_recon_path,
        output_path,
        client_name=None,
        month=None,
        first_week_start=None,
        first_week_end=None,
        last_week_start=None,
        last_week_end=None,
        bank_file_path=None,
        progress_callback=None
):
    try:
        folder = Path(invoice_folder_path)

        # Load template workbook once without saving early
        recon = openpyxl.load_workbook(template_recon_path)
        print("Template loaded with all images and formatting preserved.")

        clear_all_D_sheets(recon)
        invoice_files = list(folder.glob("*.xlsx"))
        invoices = []
        for fp in invoice_files:
            filename = Path(fp).name
            plat = detect_platform(fp)
            if plat == "Swiggy":
                d = extract_swiggy_start_day(fp)
                if d is not None:
                    invoices.append((d, fp, plat))
                    print(f"✅ Accepted Swiggy invoice: {filename} with start day {d}")
                else:
                    print(f"❌ Skipping {filename}: Could not parse start day")
            else:
                print(f"❌ Skipping {filename}: Platform not detected as Swiggy (Detected: {plat})")
        if not invoices:
            return {
                'success': False,
                'message': 'No valid Swiggy invoices found. Please check your files.'
            }

        if None in [first_week_start, first_week_end, last_week_start, last_week_end]:
            return {
                'success': False,
                'message': 'Week ranges must be provided.'
            }

        week_ranges = generate_week_ranges(first_week_start, first_week_end, last_week_start, last_week_end)

        def find_week_for_day(day):
            print(f"   🔍 Trying to map day {day} to week ranges: {week_ranges}")
            for idx, (low, high) in enumerate(week_ranges, start=1):
                if low <= high:
                    if low <= day <= high:
                        print(f"   ✅ Day {day} falls into Week {idx} ({low} to {high})")
                        return idx
                else:
                    if day >= low or day <= high:
                        print(f"   ✅ Day {day} falls into spillover Week {idx} ({low} to {high})")
                        return idx
            print(f"   ❌ Day {day} did not match any week ranges")
            return None

        week_map, summary_sheet = {}, ensure_sheet(recon, "Summary")
        if client_name:
            summary_sheet.cell(row=1, column=2).value = client_name

        col_start, row_for_weeks = 3, 4
        for i, (w_start, w_end) in enumerate(week_ranges):
            cell_letter = chr(64 + col_start + i)
            label = format_week_label(w_start, w_end)
            summary_sheet[f"{cell_letter}{row_for_weeks}"] = label

        week_expected_map = {}
        week_complaints_map = {}

        total_invoices = len(invoices)
        for idx, (d, fp, plat) in enumerate(invoices):
            if progress_callback:
                # Progress from 10% to 90%
                percent = 10 + int((idx / total_invoices) * 80)
                progress_callback(percent)

            week = find_week_for_day(d)
            if week is None:
                print(f"Invoice {fp} with start day {d} did not match any defined week range!")
                continue
            week_map[fp] = week
            print(f"\nProcessing {fp} → Week {week}")
            
            try:
                wb_invoice = openpyxl.load_workbook(fp, data_only=True, read_only=True)
            except Exception:
                wb_invoice = openpyxl.load_workbook(fp, data_only=True)
            
            # Extract Expected Receipt from Summary sheet directly
            expected_receipt = None
            try:
                sheet_sum = wb_invoice["Summary"]
                val = sheet_sum["C14"].value
                if isinstance(val, str):
                    val = val.replace(",", "").replace("₹", "").strip()
                    try:
                        val = float(val)
                    except: pass
                if isinstance(val, (int, float)):
                    expected_receipt = float(val)
            except Exception as e:
                print(f"Error extracting expected receipts from {fp}: {e}")

            if expected_receipt is not None:
                week_expected_map[week] = expected_receipt
                print(f"Invoice {fp} Week {week}: Extracted expected receipt = {expected_receipt}")
                
            # Copy data sheets
            ol = wb_invoice["Order Level"]
            d1, d2 = ensure_sheet(recon, f"D1W{week}"), ensure_sheet(recon, f"D2W{week}")
            copy_data(ol, d1, 3)
            
            add_sheet_found = False
            for sn in wb_invoice.sheetnames:
                if "Other charges and deductions" in sn or "Growth Investments" in sn:
                    add = wb_invoice[sn]
                    copy_data(add, d2, 4)
                    print(f"✅ Copied deductions sheet: '{sn}'")
                    add_sheet_found = True
                    break
            
            if not add_sheet_found:
                print(f"⚠️ Deductions/Growth sheet not found in {fp}, proceeding without it.")
            
            # Extract Total Orders directly using robust dynamic search
            total_orders = None
            try:
                sheet_sum = wb_invoice["Summary"]
                for r in range(1, 100):
                    val_b = str(sheet_sum.cell(row=r, column=2).value or "").strip().lower()
                    if "total orders (delivered + cancelled)" in val_b:
                        total_orders = sheet_sum.cell(row=r, column=3).value
                        break
            except Exception as e:
                print(f"Error extracting Total Orders from {fp}: {e}")

            wb_invoice.close()
            
            perform_calculations_on_data1(recon, d1, week, output_path)

            complaint_count = count_non_zero_complaints(d1)
            week_complaints_map[week] = complaint_count
            print(f"Week {week}: {complaint_count} non-zero complaints found")

            if total_orders is not None:
                target_col = 2 + week
                cell = summary_sheet.cell(row=6, column=target_col)
                cell.value = int(total_orders) if isinstance(total_orders, (int, float)) else total_orders
                cell.number_format = '0'
                print(f"Total Orders ({total_orders}) pasted in Summary sheet, Week {week}")
            else:
                print(f"Warning: Could not extract Total Orders from {fp}")

        for week, complaint_count in week_complaints_map.items():
            col_index = 2 + week
            cell = summary_sheet.cell(row=12, column=col_index)
            cell.value = complaint_count
            cell.number_format = '0'
            print(f"Week {week}: Complaints count {complaint_count} pasted in Summary C12 onwards")

        summary_sheet = recon["Summary"]
        convert_summary_row_to_numbers(summary_sheet, row=6, start_col='C', end_col='G')

        bank_wb = None
        try:
            if bank_file_path:
                bank_wb = openpyxl.load_workbook(bank_file_path, data_only=False)
                copy_bank_sheet_to_recon(bank_wb, recon)
                print("Bank sheet imported into reconciliation file.")
                map_bank_to_actual_receipts_from_invoice_summary(recon, week_expected_map, tolerance=10)
        except Exception as bank_e:
            print(f"Failed to process bank file: {bank_e}")
        finally:
            if bank_wb is not None:
                bank_wb.close()

        if month:
            replace_month_in_sheets(recon, month)

        add_notepoints_based_on_bank(recon, week_ranges, bank_file_path)

        # Optional: Call this to ensure images from template copied (if necessary)
        # copy_images_from_template(template_recon_path, output_path)

        recon.save(output_path)

        if bank_file_path and os.path.exists(bank_file_path):
            safe_delete_bank_file(bank_file_path, retries=10, wait=0.5)

        return {'success': True, 'message': 'Processed successfully with bank mapping, complaints count, and number formatting.'}

    except Exception as e:
        print(f"Error during processing: {str(e)}")
        import traceback
        traceback.print_exc()
        return {'success': False, 'message': f'Processing error: {str(e)}'}
