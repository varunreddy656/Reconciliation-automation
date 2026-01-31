import openpyxl
import os
import shutil
import gc
from pathlib import Path
from datetime import datetime
import calendar

# Import helpers from process_invoices
from process_invoices import (
    calculate_week_structure, 
    ensure_sheet, 
    clear_all_D_sheets, 
    perform_calculations_on_data1, 
    map_values_to_cashflow, 
    count_total_orders_from_d1w, 
    count_nonzero_compensation, 
    map_commissionable_value_to_summary,
    replace_month_in_sheets,
    parse
)
import re

def safe_float(val):
    if val is None: return 0.0
    try:
        if isinstance(val, str):
            val = val.replace('₹', '').replace(',', '').strip()
            if not val or val == '-': return 0.0
        return float(val)
    except:
        return 0.0

def parse_deduction_period(period_str):
    """Parses '01 December 25 - 07 December 25' into start and end dates"""
    try:
        parts = re.findall(r'(\d{1,2})\s+([A-Za-z]+)\s+(\d{2,4})', str(period_str))
        if len(parts) >= 2:
            def to_dt(p):
                d, m_str, y = p
                # Handle 2-digit years
                year = int(y)
                if year < 100: year += 2000
                m_num = datetime.strptime(m_str[:3].capitalize(), "%b").month
                return datetime(year, m_num, int(d)).date()
            
            return to_dt(parts[0]), to_dt(parts[1])
    except:
        pass
    return None, None

def copy_data_with_date_range(src, tgt, start_row, start_date, end_date):
    """
    Copies rows from src to tgt if 'Order Date' falls within [start_date, end_date].
    Assumes header is at start_row.
    """
    order_date_col = None
    for col_num in range(1, src.max_column + 1):
        header = str(src.cell(row=start_row, column=col_num).value or "").strip().lower()
        if header == "order date":
            order_date_col = col_num
            break

    if not order_date_col:
        print("  ⚠️ Order Date column not found in consolidated file")
        return 0

    # Write headers
    for c in range(1, src.max_column + 1):
        tgt.cell(row=1, column=c).value = src.cell(row=start_row, column=c).value

    tgt_row = 2
    copied_count = 0
    data_start_row = start_row + 1
    
    # Standardize range to date objects
    s_date = start_date.date() if isinstance(start_date, datetime) else start_date
    e_date = end_date.date() if isinstance(end_date, datetime) else end_date

    for row_values in src.iter_rows(min_row=data_start_row, values_only=True):
        try:
            date_raw = row_values[order_date_col - 1]
            if not date_raw or date_raw == '#REF!': continue
            
            row_date_dt = parse(date_raw)
            row_date = row_date_dt.date()
            
            if s_date <= row_date <= e_date:
                for col_idx, val in enumerate(row_values, 1):
                    tgt.cell(row=tgt_row, column=col_idx).value = val
                tgt_row += 1
                copied_count += 1
        except:
            continue
            
    return copied_count

def process_zomato_consolidated(
    invoice_folder,
    template_path,
    output_path,
    client_name=None,
    month="October",
    first_week_start=None,
    first_week_end=None,
    last_week_start=None,
    last_week_end=None,
    progress_callback=None
):
    """
    Consolidated Zomato Reconciliation Logic.
    Splits one monthly file into weekly subsheets based on user-provided week ranges.
    """
    try:
        if progress_callback: progress_callback(5)

        # 1. Setup Template
        print(f"🔄 Copying template: {template_path} → {output_path}")
        shutil.copy2(template_path, output_path)
        recon = openpyxl.load_workbook(output_path)
        clear_all_D_sheets(recon)
        
        if progress_callback: progress_callback(15)

        # 2. Find the Consolidated File
        folder = Path(invoice_folder)
        files = list(folder.glob("*.xlsx"))
        if not files:
            return {'success': False, 'message': 'No consolidated file found'}
        
        consolidated_fp = files[0]
        print(f"📊 Processing Consolidated File: {consolidated_fp.name}")
        wb_source = openpyxl.load_workbook(consolidated_fp, data_only=True, read_only=True)
        
        # Find Order level sheet
        possible_order_sheets = ["Order Level", "Order level", "Order Details", "Orders"]
        src_ol = None
        for sn in possible_order_sheets:
            if sn in wb_source.sheetnames:
                src_ol = wb_source[sn]
                break
        
        if not src_ol:
            return {'success': False, 'message': 'Order Level sheet not found in consolidated file'}

        # 3. Calculate Week Structure
        week_structure = calculate_week_structure(
            month, first_week_start, first_week_end, last_week_start, last_week_end
        )
        total_weeks = len(week_structure)
        print(f"📅 Identified {total_weeks} weeks")

        summary_sheet = ensure_sheet(recon, "Summary")
        if client_name:
            summary_sheet.cell(row=1, column=2).value = client_name

        # 4. Process Each Week
        for idx, week in enumerate(week_structure):
            if progress_callback:
                progress_callback(20 + int((idx/total_weeks) * 70))
            
            week_num = week['week_num']
            print(f"\n--- Processing Week {week_num}: {week['label']} ---")
            
            # Create D1W sheet
            d1 = ensure_sheet(recon, f"D1W{week_num}")
            
            # Filter and copy data for this week range
            copied = copy_data_with_date_range(src_ol, d1, 7, week['start_date'], week['end_date'])
            print(f"  ✅ Extracted {copied} rows")
            
            if copied > 0:
                # Perform standard Zomato calculations and mapping
                perform_calculations_on_data1(recon, d1, week_num, output_path)
                map_values_to_cashflow(recon, d1, week_num)
                
                # Update Summary
                summary_col = 3 + (week_num - 1)
                summary_sheet.cell(row=4, column=summary_col).value = week['label']
                
                total_orders = count_total_orders_from_d1w(d1, header_row=5)
                summary_sheet.cell(row=6, column=summary_col).value = total_orders
                
                comp_orders = count_nonzero_compensation(d1, data_row=5)
                summary_sheet.cell(row=12, column=summary_col).value = comp_orders
                
                map_commissionable_value_to_summary(summary_sheet, d1, week_num)

                map_commissionable_value_to_summary(summary_sheet, d1, week_num)
        
        # 5. Process Ads Segregation
        print("\n📢 Processing Ads Segregation...")
        ads_weekly_totals = {w['week_num']: 0.0 for w in week_structure}
        
        src_ads = None
        for sn in ["Addition Deductions Details", "Addition Deductions", "Deductions"]:
            if sn in wb_source.sheetnames:
                src_ads = wb_source[sn]
                break
        
        if src_ads:
            # Find the starting row of sections
            current_section = None
            type_col, period_col, total_col = -1, -1, -1
            
            for row in range(1, src_ads.max_row + 1):
                row_val_b = str(src_ads.cell(row=row, column=2).value or "").strip().lower()
                
                # Check for section headers
                if "addition type" in row_val_b:
                    current_section = "ADDITION"
                    continue
                elif "deduction type" in row_val_b:
                    current_section = "DEDUCTION"
                    continue
                elif "investments in hyperpure" in row_val_b or "other deductions" in row_val_b:
                    current_section = "OTHER"
                    continue

                # Look for column headers if not found yet
                if type_col == -1:
                    headers = [str(src_ads.cell(row=row, column=c).value or "").strip().lower() for c in range(1, 15)]
                    for idx, h in enumerate(headers, 1):
                        if "type" == h: type_col = idx
                        if "deduction time period" in h or "order date" in h: period_col = idx
                        if "total amount" in h: total_col = idx
                    continue
                
                # Process ADS rows
                if current_section in ["ADDITION", "DEDUCTION"] and type_col != -1:
                    type_val = str(src_ads.cell(row=row, column=type_col).value or "").strip().upper()
                    if type_val == "ADS":
                        period_val = src_ads.cell(row=row, column=period_col).value
                        amount = safe_float(src_ads.cell(row=row, column=total_col).value)
                        
                        # Use deduction period if available, else try to find any date in the row
                        p_start, p_end = parse_deduction_period(period_val)
                        if not p_start:
                            # Fallback: maybe it's a simple date
                            try:
                                dt = parse(period_val).date()
                                p_start, p_end = dt, dt
                            except: pass

                        if p_start:
                            # Match to week
                            matched = False
                            for week in week_structure:
                                ws, we = week['start_date'].date(), week['end_date'].date()
                                if ws <= p_start <= we or (p_start <= ws and p_end >= we):
                                    # DEDUCT if in Addition section, ADD if in Deduction section
                                    if current_section == "ADDITION":
                                        ads_weekly_totals[week['week_num']] -= amount
                                        print(f"  ➖ Deducted Addition Ad: {amount} for Week {week['week_num']}")
                                    else:
                                        ads_weekly_totals[week['week_num']] += amount
                                        print(f"  ➕ Added Deduction Ad: {amount} for Week {week['week_num']}")
                                    matched = True
                                    break
                            if not matched:
                                print(f"  ⚠️ Could not match Ad period '{period_val}' to any week")
            
            
        else:
            print("  ⚠️ 'Addition Deductions Details' sheet not found")

        # 6. Check D1W sheets for Extra inventory ads (always check)
        print("\n🔍 Checking D1W sheets for 'Extra inventory ads (order level deduction)'...")
        for week in week_structure:
            wn = week['week_num']
            d1_name = f"D1W{wn}"
            if d1_name in recon.sheetnames:
                d1_sheet = recon[d1_name]
                # Search row 5 for header
                for col_idx in range(1, d1_sheet.max_column + 1):
                    h = str(d1_sheet.cell(row=5, column=col_idx).value or "").strip().lower()
                    if "extra inventory ads (order level deduction)" in h:
                        val = d1_sheet.cell(row=4, column=col_idx).value
                        if isinstance(val, (int, float)):
                            ads_weekly_totals[wn] += val
                            print(f"  ✅ Week {wn} Extra Ads: {val} (Total Ads now: {ads_weekly_totals[wn]})")
                        break

        # 7. Map Ads to Cashflow
        if "Cashflow" in recon.sheetnames:
            cashflow = recon["Cashflow"]
            high_priority_row = -1
            for r in range(1, cashflow.max_row + 1):
                label = str(cashflow.cell(row=r, column=2).value or "").strip().lower()
                if label == "high priority":
                    high_priority_row = r
                    break
            
            if high_priority_row != -1:
                for week_num, total in ads_weekly_totals.items():
                    col = 3 + (week_num - 1)
                    # For consolidated, we set the value directly
                    cashflow.cell(row=high_priority_row, column=col).value = total
                    print(f"  ✅ Week {week_num} Final Ads: {total} → Cashflow[{high_priority_row}, {col}]")
            else:
                print("  ⚠️ 'High Priority' label not found in Cashflow sheet")

        # 7. Finalize
        replace_month_in_sheets(recon, month)
        recon.save(output_path)
        recon.close()
        wb_source.close()
        gc.collect()

        if progress_callback: progress_callback(100)
        return {
            'success': True,
            'message': f'Consolidated processing complete! Split into {total_weeks} weeks.',
            'weeks_processed': total_weeks
        }

    except Exception as e:
        import traceback
        traceback.print_exc()
        return {'success': False, 'message': f'Error: {str(e)}'}
