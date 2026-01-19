from openpyxl import load_workbook
import openpyxl
from pathlib import Path
import re
import shutil
import os
import time
import tkinter as tk
from tkinter import filedialog, messagebox
from datetime import datetime, timedelta
import tempfile
import gc  # Add at top
import calendar
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


# ===================== ZOMATO-SPECIFIC HELPERS =====================

def parse(date_str, dayfirst=True):
    """Custom date parser - replacement for dateutil.parser.parse"""
    # Handle integer input (just day number)
    if isinstance(date_str, (int, float)):
        return datetime(2025, 1, int(date_str))

    date_str = str(date_str).strip()

    # Handle simple day number string "27", "2", etc.
    if date_str.isdigit():
        return datetime(2025, 1, int(date_str))

    # Try common date formats
    formats = [
        "%d/%m/%Y",
        "%d-%m-%Y",
        "%Y-%m-%d",
        "%d %B %Y",
        "%d %b %Y",
        "%d.%m.%Y"
    ]

    for fmt in formats:
        try:
            return datetime.strptime(date_str, fmt)
        except:
            continue

    # Last resort - extract first number
    import re
    match = re.search(r'\d+', date_str)
    if match:
        return datetime(2025, 1, int(match.group()))

    raise ValueError(f"Cannot parse date: {date_str}")


def calculate_week_structure(month, first_week_start, first_week_end, last_week_start, last_week_end):
    """
    Calculate complete week structure with spillover handling

    Returns: List of dicts with week info
    """

    year = 2025  # Current year

    # Get month number from name
    month_num = datetime.strptime(month, "%B").month

    # Parse day numbers (they come as strings like "27", "2", etc.)
    first_start_day = int(str(first_week_start).strip())
    first_end_day = int(str(first_week_end).strip())
    last_start_day = int(str(last_week_start).strip())
    last_end_day = int(str(last_week_end).strip())

    # Determine actual months for each date
    # If first week start day > first week end day, it's from previous month
    if first_start_day > first_end_day:
        first_start_month = month_num - 1 if month_num > 1 else 12
        first_start_year = year if month_num > 1 else year - 1
    else:
        first_start_month = month_num
        first_start_year = year

    first_end_month = month_num
    first_end_year = year

    last_start_month = month_num
    last_start_year = year

    # If last week end day < last week start day, it's in next month
    if last_end_day < last_start_day:
        last_end_month = month_num + 1 if month_num < 12 else 1
        last_end_year = year if month_num < 12 else year + 1
    else:
        last_end_month = month_num
        last_end_year = year

    # Create datetime objects
    first_start = datetime(first_start_year, first_start_month, first_start_day)
    first_end = datetime(first_end_year, first_end_month, first_end_day)
    last_start = datetime(last_start_year, last_start_month, last_start_day)
    last_end = datetime(last_end_year, last_end_month, last_end_day)

    weeks = []

    # ✅ FIXED: First week - only show dates in target month for label
    # For label: if spillover at start, show from 1st of target month
    first_week_label_start = 1 if first_start.month != month_num else first_start.day
    first_week_label_end = first_end.day

    week_1 = {
        'week_num': 1,
        'start_date': first_start,
        'end_date': first_end,
        'label': f"{ordinal(first_week_label_start)} to {ordinal(first_week_label_end)}",
        'is_spillover_start': first_start.month != month_num,
        'is_spillover_end': False
    }
    weeks.append(week_1)

    # Calculate middle weeks (always 7 days)
    current_start = first_end + timedelta(days=1)
    week_counter = 2

    while current_start < last_start:
        current_end = current_start + timedelta(days=6)

        week = {
            'week_num': week_counter,
            'start_date': current_start,
            'end_date': current_end,
            'label': f"{ordinal(current_start.day)} to {ordinal(current_end.day)}",
            'is_spillover_start': False,
            'is_spillover_end': False
        }
        weeks.append(week)

        current_start = current_end + timedelta(days=1)
        week_counter += 1

    # ✅ FIXED: Last week - only show dates in target month for label
    # For label: if spillover at end, show till last day of target month
    last_week_label_start = last_start.day
    last_week_label_end = calendar.monthrange(year, month_num)[1] if last_end.month != month_num else last_end.day

    week_last = {
        'week_num': week_counter,
        'start_date': last_start,
        'end_date': last_end,
        'label': f"{ordinal(last_week_label_start)} to {ordinal(last_week_label_end)}",
        'is_spillover_start': False,
        'is_spillover_end': last_end.month != month_num
    }
    weeks.append(week_last)

    return weeks


def match_invoice_to_week(invoice_filename, week_structure, month):
    """
    Match invoice filename to correct week in structure
    Returns: week_num or None if no match
    """

    # Parse filename dates
    # Expected format: "Restaurant_Name_27_Oct_to_2_Nov_2024.xlsx"
    # OR: "27_Oct_2024_2_Nov_2024.xlsx"

    try:
        # Remove .xlsx/.xls extension
        name_without_ext = invoice_filename.replace('.xlsx', '').replace('.xls', '')

        # Split by underscore
        parts = name_without_ext.split('_')

        print(f"\n🔍 Parsing invoice: {invoice_filename}")
        print(f"   Parts: {parts}")

        # Try to find date pattern: DAY_MONTH_...DAY_MONTH_YEAR
        # Look for numbers followed by month names

        start_day = None
        start_month = None
        end_day = None
        end_month = None
        year = 2025  # Default year

        # Find first date (start)
        for i in range(len(parts) - 1):
            if parts[i].isdigit() and len(parts[i]) <= 2:
                # Check if next part is a month
                if i + 1 < len(parts) and parts[i + 1][:3].capitalize() in [
                    'Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
                    'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec'
                ]:
                    start_day = int(parts[i])
                    start_month = parts[i + 1][:3].capitalize()

                    # Look for year after this
                    if i + 2 < len(parts) and parts[i + 2].isdigit() and len(parts[i + 2]) == 4:
                        year = int(parts[i + 2])

                    # Now find end date (continue from current position)
                    for j in range(i + 2, len(parts) - 1):
                        if parts[j].isdigit() and len(parts[j]) <= 2:
                            if j + 1 < len(parts) and parts[j + 1][:3].capitalize() in [
                                'Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
                                'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec'
                            ]:
                                end_day = int(parts[j])
                                end_month = parts[j + 1][:3].capitalize()

                                # Check for year after end date
                                if j + 2 < len(parts) and parts[j + 2].isdigit() and len(parts[j + 2]) == 4:
                                    year = int(parts[j + 2])
                                break
                    break

        if not all([start_day, start_month, end_day, end_month]):
            print(f"   ❌ Could not parse dates from filename")
            return None

        print(f"   ✅ Parsed: {start_day} {start_month} to {end_day} {end_month} {year}")

        # Convert month names to numbers
        month_map = {
            'Jan': 1, 'Feb': 2, 'Mar': 3, 'Apr': 4, 'May': 5, 'Jun': 6,
            'Jul': 7, 'Aug': 8, 'Sep': 9, 'Oct': 10, 'Nov': 11, 'Dec': 12
        }

        start_month_num = month_map[start_month]
        end_month_num = month_map[end_month]

        # Create datetime objects
        invoice_start = datetime(year, start_month_num, start_day)
        invoice_end = datetime(year, end_month_num, end_day)

    except Exception as e:
        print(f"   ❌ Error parsing invoice filename: {e}")
        import traceback
        traceback.print_exc()
        return None

    # Match to week structure
    for week in week_structure:
        if (week['start_date'].date() == invoice_start.date() and
                week['end_date'].date() == invoice_end.date()):
            print(f"   ✅ MATCHED to Week {week['week_num']}")
            return week['week_num']

    print(f"   ⚠️ No matching week found")
    print(f"      Invoice dates: {invoice_start.date()} to {invoice_end.date()}")
    print(f"      Available weeks:")
    for week in week_structure:
        print(f"         Week {week['week_num']}: {week['start_date'].date()} to {week['end_date'].date()}")
    return None


def ordinal(n):
    """Convert number to ordinal (1 -> 1st, 2 -> 2nd, etc.)"""
    if 10 <= n % 100 <= 20:
        suffix = 'th'
    else:
        suffix = {1: 'st', 2: 'nd', 3: 'rd'}.get(n % 10, 'th')
    return f"{n}{suffix}"


def extract_zomato_week_range(filepath):
    """Extract week range PRIORITIZED from FILENAME (most reliable)"""
    filename = Path(filepath).name

    # 🎯 PRIMARY: Parse filename like "01_Sep_2025_07_Sep_2025"
    # ✅ FIXED REGEX - Don't capture year as day!
    date_match = re.search(r'(\d{1,2})_([A-Za-z]{3})_\d{4}_(\d{1,2})_([A-Za-z]{3})_\d{4}', filename)

    if date_match:
        start_day = int(date_match.group(1))
        start_month = date_match.group(2)
        end_day = int(date_match.group(3))
        end_month = date_match.group(4)

        print(f"✅ FILENAME PARSED: {start_day} {start_month} → {end_day} {end_month}")

        return {
            'start_day': start_day,
            'start_month': start_month,
            'end_day': end_day,
            'end_month': end_month,
            'full_text': f"{start_day} {start_month} to {end_day} {end_month}"
        }

    # 🔍 FALLBACK: Try HSummary!E2 or Summary!C4
    wb = None
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True)
        try:
            hsummary = wb["HSummary"]
            text = str(hsummary["E2"].value or "")
        except:
            summary = wb["Summary"]
            text = str(summary["C4"].value or "")

        # Parse text patterns
        patterns = [
            r'(\d{1,2})\s*([A-Za-z]{3})\s*[-to]+\s*(\d{1,2})\s*([A-Za-z]{3})',
            r'(\d{1,2})\s*-\s*(\d{1,2})\s*([A-Za-z]{3})',
        ]

        for pattern in patterns:
            m = re.search(pattern, text, re.IGNORECASE)
            if m and len(m.groups()) >= 2:
                start_day = int(m.group(1))
                end_day = int(m.group(2)) if len(m.groups()) >= 3 else int(m.group(1))
                month = m.group(3)[:3] if len(m.groups()) > 2 else m.group(2)[:3]
                print(f"✅ SHEET PARSED: {start_day}-{end_day} {month}")
                return {
                    'start_day': start_day,
                    'start_month': month,
                    'end_day': end_day,
                    'end_month': month,
                    'full_text': f"{start_day} {month} to {end_day} {month}"
                }
    except:
        pass
    finally:
        if wb:
            wb.close()

    print(f"❌ Could not parse filename or sheets: {filename}")
    return None


def select_invoices_gui():
    """GUI: ONLY invoices + template (auto temp folder + output)"""
    root = tk.Tk()
    root.withdraw()

    # 1️⃣ Select multiple invoices ONLY
    invoice_paths = filedialog.askopenfilenames(
        title="Select Zomato Invoice Files (hold Ctrl/Cmd for multiple)",
        filetypes=[("Excel files", "*.xlsx *.xls")]
    )

    if not invoice_paths:
        return None

    # 2️⃣ Select template ONLY
    template_path = filedialog.askopenfilename(
        title="Select Reconciliation Template",
        filetypes=[("Excel files", "*.xlsx *.xls")]
    )

    if not template_path:
        return None

    # 🔄 AUTO: Create temp folder + copy invoices
    invoice_folder_path = tempfile.mkdtemp(prefix="zomato_invoices_")
    copied_files = []

    for fp in invoice_paths:
        filename = os.path.basename(fp)
        dest = os.path.join(invoice_folder_path, filename)
        shutil.copy2(fp, dest)
        copied_files.append(dest)
        print(f"✅ Copied: {filename}")

    # 🔄 AUTO: Generate output path (template name + _output)
    output_name = os.path.splitext(os.path.basename(template_path))[0] + "_output.xlsx"
    output_folder = os.path.dirname(template_path)
    output_path = os.path.join(output_folder, output_name)

    print(f"📁 Temp invoices: {invoice_folder_path}")
    print(f"💾 Output will be: {output_path}")

    return {
        'invoice_folder_path': invoice_folder_path,
        'template_recon_path': template_path,
        'output_path': output_path,
        'copied_files': copied_files
    }


def parse_month_to_days(month_name):
    """Convert month name to (start_day=1, end_day) for that month"""
    days_in_month = {
        "January": 31, "February": 28, "March": 31, "April": 30, "May": 31,
        "June": 30, "July": 31, "August": 31, "September": 30, "October": 31,
        "November": 30, "December": 31
    }
    return 1, days_in_month.get(month_name, 31)


def month_str_to_num(month_str):
    """Convert 'Oct', 'Sep' to month number"""
    months = {
        "Jan": 1, "Feb": 2, "Mar": 3, "Apr": 4, "May": 5, "Jun": 6,
        "Jul": 7, "Aug": 8, "Sep": 9, "Oct": 10, "Nov": 11, "Dec": 12
    }
    return months.get(month_str, None)


def split_invoice_by_month(week_range, target_month_name):
    """ALWAYS include invoices if they have target month days"""
    if not week_range:
        return None

    start_day = week_range['start_day']
    start_month = week_range['start_month']
    end_day = week_range['end_day']
    end_month = week_range['end_month']

    target_month_num = month_str_to_num(target_month_name[:3])
    start_month_num = month_str_to_num(start_month[:3])
    end_month_num = month_str_to_num(end_month[:3])

    target_month_start, target_month_end = parse_month_to_days(target_month_name)

    parts = {
        'opening_adj': None, 'normal_week': None, 'closing_adj': None, 'week_label': None
    }

    invoice_overlaps = False

    # Case 1: Invoice is entirely within target month
    if start_month_num == target_month_num and end_month_num == target_month_num:
        invoice_overlaps = True
        normal_start = start_day
        normal_end = end_day

    # Case 2: Invoice starts before and ends in target month
    elif start_month_num < target_month_num and end_month_num == target_month_num:
        invoice_overlaps = True
        normal_start = 1
        normal_end = end_day
        parts['opening_adj'] = (start_day, target_month_start - 1)
        print(f"📅 OPENING spillover: {start_day} {start_month}")

    # Case 3: Invoice starts in target month and ends after
    elif start_month_num == target_month_num and end_month_num > target_month_num:
        invoice_overlaps = True
        normal_start = start_day
        normal_end = target_month_end
        parts['closing_adj'] = (target_month_end + 1, end_day)
        print(f"📅 CLOSING spillover: {end_day} {end_month}")

    # Case 4: Invoice spans across target month
    elif start_month_num < target_month_num and end_month_num > target_month_num:
        invoice_overlaps = True
        normal_start = 1
        normal_end = target_month_end
        parts['opening_adj'] = (start_day, target_month_start - 1)
        parts['closing_adj'] = (target_month_end + 1, end_day)
        print(f"📅 OPENING spillover: {start_day} {start_month}, CLOSING: {end_day} {end_month}")

    if invoice_overlaps:
        parts['normal_week'] = (normal_start, normal_end)
        parts['week_label'] = f"{ordinal(normal_start)} to {ordinal(normal_end)}"
        print(f"✅ Week created: {parts['week_label']} (Month: {target_month_name})")
        return parts
    else:
        print(f"⚠️  Invoice does not overlap with {target_month_name}")
        return None


def ordinal(n):
    """Convert number to ordinal (1→1st, 2→2nd, 3→3rd, etc.)"""
    if 11 <= (n % 100) <= 13:
        return f"{n}th"
    suffixes = {1: "st", 2: "nd", 3: "rd"}
    return f"{n}{suffixes.get(n % 10, 'th')}"


def clear_all_D_sheets(wb):
    """Remove old D1W and D2W sheets"""
    sheets_to_remove = [sh for sh in wb.sheetnames if sh.startswith("D1W") or sh.startswith("D2W")]
    for sh_name in sheets_to_remove:
        std = wb[sh_name]
        wb.remove(std)
    print(f"Cleared {len(sheets_to_remove)} old D1W/D2W sheets.")


def ensure_sheet(wb, name):
    """Get or create sheet"""
    if name in wb.sheetnames:
        return wb[name]
    else:
        return wb.create_sheet(name)


def copy_data_with_spillover_filter(src, tgt, start_row, target_month=None, week_info=None, cashflow_sheet=None):
    """Smart copy: Filter by target month + calculate spillover adjustments."""

    order_date_col = None
    payout_col = None

    for col_num in range(1, src.max_column + 1):
        header = str(src.cell(row=start_row, column=col_num).value or "").strip()
        if header.lower() == "order date":
            order_date_col = col_num
        if "order level payout" in header.lower():
            payout_col = col_num

    if not order_date_col or not target_month:
        print(f"  ⚠️  No spillover filtering - copying all data")
        max_row, max_col = src.max_row, src.max_column
        tgt.delete_rows(1, tgt.max_row)
        for r in range(start_row, max_row + 1):
            for c in range(1, max_col + 1):
                tgt.cell(row=r - start_row + 1, column=c).value = src.cell(row=r, column=c).value
        print(f"  📊 Copied {max_row - start_row + 1} rows")
        return None

    print(f"  ✅ Order Date column: {order_date_col}")
    if payout_col:
        print(f"  ✅ Order level Payout column: {payout_col}")

    target_month_num = month_str_to_num(target_month[:3])

    opening_spillover_sum = 0
    closing_spillover_sum = 0

    tgt.delete_rows(1, tgt.max_row)
    for c in range(1, src.max_column + 1):
        tgt.cell(row=1, column=c).value = src.cell(row=start_row, column=c).value

    tgt_row = 2
    copied_rows = 0
    opening_rows = 0
    closing_rows = 0
    max_col = src.max_column
    data_start_row = start_row + 1

    print(f"  🔄 Scanning data rows {data_start_row} to {src.max_row}...")

    for src_row in range(data_start_row, src.max_row + 1):
        date_value = src.cell(row=src_row, column=order_date_col).value

        row_month_num = None

        if isinstance(date_value, datetime):
            row_month_num = date_value.month
        elif isinstance(date_value, str):
            date_str = date_value.strip()
            if date_str == '#REF!' or not date_str:
                continue
            date_match = re.match(r'(\d{4})-(\d{2})-(\d{2})', date_str)
            if date_match:
                row_month_num = int(date_match.group(2))

        if not row_month_num:
            continue

        payout_value = 0
        if payout_col:
            payout_value = src.cell(row=src_row, column=payout_col).value
            if not isinstance(payout_value, (int, float)):
                payout_value = 0

        if row_month_num < target_month_num:
            opening_spillover_sum += payout_value
            opening_rows += 1
        elif row_month_num > target_month_num:
            closing_spillover_sum += payout_value
            closing_rows += 1
        else:
            for c in range(1, max_col + 1):
                tgt.cell(row=tgt_row, column=c).value = src.cell(row=src_row, column=c).value
            tgt_row += 1
            copied_rows += 1

    print(f"  📊 Copied {copied_rows} data rows (target month: {target_month})")
    print(f"  📊 Opening spillover: {opening_rows} rows → Sum: {opening_spillover_sum}")
    print(f"  📊 Closing spillover: {closing_rows} rows → Sum: {closing_spillover_sum}")

    return {
        'opening_spillover': opening_spillover_sum,
        'closing_spillover': closing_spillover_sum,
        'week_num': week_info['week_num'] if week_info else None
    }


def count_total_orders_from_d1w(d1_sheet, header_row=5):
    """
    Count ALL orders - simply count all rows below header
    """
    try:
        data_start_row = header_row + 1  # Row 6 onwards

        # Find the ACTUAL last row with data
        last_data_row = data_start_row

        for row_num in range(data_start_row, d1_sheet.max_row + 1):
            # Check first 5 columns for any data
            has_data = False
            for col in range(1, 6):  # Check columns A-E
                cell_value = d1_sheet.cell(row=row_num, column=col).value
                if cell_value is not None and str(cell_value).strip() != '':
                    has_data = True
                    last_data_row = row_num
                    break

        # Count from data_start to last_data_row
        count = last_data_row - data_start_row + 1

        # Sanity check - if count is negative or zero, return 0
        if count <= 0:
            print(f"  ⚠️  No data rows found after header row {header_row}")
            return 0

        print(f"  📊 Total Orders: {count} (rows {data_start_row} to {last_data_row})")
        return count

    except Exception as e:
        print(f"  ❌ Error counting orders: {e}")
        import traceback
        traceback.print_exc()
        return 0


def count_nonzero_compensation(sheet, data_row=1):
    """
    Count non-zero values in 'Customer Compensation/Recoupment' column
    data_row should be the row with calculated values (row 1 after 4 rows inserted)
    """
    try:
        # Find Customer Compensation/Recoupment column in HEADER ROW (row 5)
        compensation_col = None
        header_row = 5  # Header is always at row 5 after inserting 4 rows

        print(f"  🔍 Searching for compensation column in row {header_row}...")

        for col_num in range(1, sheet.max_column + 1):
            header = str(sheet.cell(row=header_row, column=col_num).value or "").strip().lower()

            if "customer compensation" in header and "recoupment" in header:
                compensation_col = col_num
                print(f"  ✅ Found compensation column: Col {col_num} - '{header}'")
                break

        if not compensation_col:
            print(f"  ❌ Compensation column not found")
            return 0

        # Get the SUM value from the calculated row (row 1, 2, or 4)
        # Row 1 = Cancelled, Row 2 = Delivered, Row 4 = Total
        # We want to count from DATA ROWS (row 6 onwards), not from calculated rows!

        count = 0
        data_start_row = header_row + 1  # Row 6 onwards

        print(f"  🔄 Counting non-zero compensation from row {data_start_row} to {sheet.max_row}...")

        for row in range(data_start_row, sheet.max_row + 1):
            value = sheet.cell(row=row, column=compensation_col).value

            # Check if value is non-zero
            if value is not None:
                try:
                    num_value = float(value)
                    if num_value != 0:
                        count += 1
                except (ValueError, TypeError):
                    pass

        print(f"  📊 Non-zero compensation count: {count}")
        return count

    except Exception as e:
        print(f"  ❌ Error counting compensation: {e}")
        import traceback
        traceback.print_exc()
        return 0


ZOMATO_MAPPING = {
    "Item sales (Delivered orders)": (["Subtotal (items total)"], 2, "single"),
    "Add:- Packing charges": (["Packaging Charge", "Packing charge"], 2, "single"),
    "Add:- Compensation paid for cancelled orders": (
    ["Net Additions \n(cancellation refund for cancelled orders/ tip for kitchen staff for delivered orders)",
     "Net Additions \n(cancellation refund for cancelled orders/Tip for Kitchen Staff for delivered orders)",
     "Net Additions \n(cancellation refund for cancelled orders)"], 1, "single"),
    "Less:- Discount": (
        ["Restaurant discount (Promo)",
         "Restaurant discount (BOGO, Freebies, Gold, Brand pack & others)",
         "Delivery charge discount/ Relisting discount ",
         "Restaurant discount [Promo]",
         "Restaurant discount (BOGO, Freebies, Gold, Brand pack & others)",
         "Delivery charge discount/ Relisting discount",
         "Restaurant Discount [BOGO, Freebies, Gold, Brand pack & others]",
         "Delivery charge discount / Relisting discount",
         "Restaurant discount [Flat offs, Freebies, Gold, Brand pack & others]"],
        2, "sum"),
    "Add:- GST": (["Total GST collected from customers"], 2, "single"),
    "Platform Fee": (["Base service fee", "Service fee\n[ (9) * (10) ]", "Base service fee\n[(12)% * (B)]","Base service fee\n[ (9) * (10) ]"], 3, "single"),
    "Convenience Fee": (["Payment mechanism fee"], 3, "single"),
    "Discount on Service Fee": (
        ["Discount on service fee due to 30% capping",
         "Discount due to 30% service fee capping (excluding GST)\n\nService fees capped at 30% of commissionable value (9)",
         "Discount on service fee due to 30% capping\n\nService fees capped at 30% of commissionable value (B)"], 3, "single"),
    "Long Distance Fee": (["Long distance enablement fee", "Fulfilment fee","Fulfilment fee\n[ (13) * (14) ]"], 3, "special"),
    "Paid by Restaurant": (
        ["Customer Compensation/Recoupment", "Customer compensation/recoupment", "Customer compensation/ recoupment"],
        4, "single"),
    "TDS deduction for aggrigators": (["TDS 194O amount"], 4, "single"),
    "TCS": (["TCS IGST amount", "Tax collected at source"], 4, "sum"),
    "GST collected and paid by Zomato": (
        ["GST paid by Zomato on behalf of restaurant - under section 9(5)"], 4, "single"),
}


def perform_calculations_on_data1(wb, data1_sheet, week, recon_path):
    """Add 4 rows + calculate with EXACT column matching"""
    data1_sheet.insert_rows(1, 4)
    print("✅ Inserted 4 rows at top")

    header_row = 5
    item_total_col = None
    order_status_col = None

    print(f"🔍 Scanning headers in row {header_row}...")
    for col_num in range(1, data1_sheet.max_column + 1):
        header = str(data1_sheet.cell(row=header_row, column=col_num).value or "")
        header_lower = header.lower()

        if header_lower == "order status (delivered/ cancelled/ rejected)":
            order_status_col = col_num
            print(f"✅ EXACT ORDER STATUS at col {order_status_col}: '{header}'")

        if header_lower == "subtotal (items total)":
            item_total_col = col_num
            print(f"✅ EXACT SUBTOTAL at col {item_total_col}: '{header}'")

    if not item_total_col:
        print("❌ 'Subtotal (items total)' NOT FOUND!")
        return

    if not order_status_col:
        print("❌ 'Order status (Delivered/ Cancelled/ Rejected)' NOT FOUND!")
        return

    print(f"📊 Summing from col {item_total_col}")

    delivered = [0] * (data1_sheet.max_column - item_total_col + 1)
    cancelled = [0] * (data1_sheet.max_column - item_total_col + 1)

    data_start_row = header_row + 1
    skipped_rows = 0

    for row in range(data_start_row, data1_sheet.max_row + 1):
        status_text = str(data1_sheet.cell(row=row, column=order_status_col).value or "").upper().strip()

        if status_text == "DELIVERED":
            status = "delivered"
        elif status_text in ["CANCELLED", "TIMEDOUT", "TIMEOUT", "REJECTED"]:
            status = "cancelled"
        else:
            skipped_rows += 1
            continue

        target = delivered if status == "delivered" else cancelled

        for i, col in enumerate(range(item_total_col, data1_sheet.max_column + 1)):
            val = data1_sheet.cell(row=row, column=col).value
            if isinstance(val, (int, float)) and val != 0:
                target[i] += val

    for i, col in enumerate(range(item_total_col, data1_sheet.max_column + 1)):
        data1_sheet.cell(row=1, column=col).value = cancelled[i]
        data1_sheet.cell(row=2, column=col).value = delivered[i]
        data1_sheet.cell(row=3, column=col).value = delivered[i] * 1.18
        data1_sheet.cell(row=4, column=col).value = delivered[i] + cancelled[i]

    if skipped_rows > 0:
        print(f"⚠️  Skipped {skipped_rows} rows (REJECTED or unknown status)")

    print("✅ EXACT MATCH - CANCELLED/DELIVERED calculations COMPLETE!")


def map_values_to_cashflow(wb, data1_sheet, week, week_type="normal"):
    """Map Zomato D1W data to Cashflow sheet"""
    if "Cashflow" not in wb.sheetnames:
        print("Cashflow sheet not found")
        return

    cashflow = wb["Cashflow"]

    if week_type == "opening_adj":
        week_col = 2
    elif week_type == "closing_adj":
        week_col = 100
    else:
        week_col = 3 + (week - 1)

    headers = {}
    for col_num in range(1, data1_sheet.max_column + 1):
        header_val = data1_sheet.cell(row=5, column=col_num).value
        if header_val:
            headers[str(header_val).strip().lower()] = col_num

    def find_all_columns(search_list):
        matches = set()
        for search_name in search_list:
            search_clean = search_name.strip().lower()
            for actual_header, col_num in headers.items():
                if actual_header.lower() == search_clean:
                    matches.add(col_num)
                    print(f"  ✅ Exact match '{search_name}' → Col {col_num}")
        return list(matches)

    print("🔍 ALL D1W HEADERS (row 5):")
    for col_num in range(1, data1_sheet.max_column + 1):
        header_raw = data1_sheet.cell(row=5, column=col_num).value
        header = str(header_raw or "").strip()
        if header:
            print(f"  Col {col_num}: '{header}'")
    print()

    for row in range(1, cashflow.max_row + 1):
        label = str(cashflow.cell(row=row, column=2).value or "").strip()
        if not label or label not in ZOMATO_MAPPING:
            continue

        header_list, data_row, operation = ZOMATO_MAPPING[label]
        all_matching_cols = set()

        for header_name in header_list:
            matching_cols = find_all_columns([header_name])
            all_matching_cols.update(matching_cols)

        all_matching_cols = list(all_matching_cols)
        data_cells = [data1_sheet.cell(row=data_row, column=col) for col in all_matching_cols]

        print(f"📊 '{label}': Found {len(data_cells)} columns → Row {data_row}")

        if not data_cells:
            print(f"❌ Skipping '{label}' - No matching columns found")
            continue

        formula = None

        if label == "Long Distance Fee":
            fee_cols = all_matching_cols
            discount_cols = find_all_columns([
                "Discount on long distance enablement fee",
                "Discount on Fulfilment fee"
            ])

            if fee_cols:
                fee_cell = data1_sheet.cell(row=data_row, column=fee_cols[0]).coordinate
                if discount_cols:
                    discount_cell = data1_sheet.cell(row=data_row, column=discount_cols[0]).coordinate
                    formula = f"='{data1_sheet.title}'!{fee_cell}-'{data1_sheet.title}'!{discount_cell}"
                    print(f"  ✅ Long Distance formula (with discount): {formula}")
                else:
                    formula = f"='{data1_sheet.title}'!{fee_cell}"
                    print(f"  ✅ Long Distance formula (no discount): {formula}")

        elif operation == "single":
            formula = f"='{data1_sheet.title}'!{data_cells[0].coordinate}"
        elif operation == "sum":
            formula = "=" + "+".join([f"'{data1_sheet.title}'!{c.coordinate}" for c in data_cells])
        elif operation == "sub":
            if len(data_cells) < 2:
                print(f"WARNING: Subtraction for '{label}' needs 2+ cells")
                continue
            formula = f"='{data1_sheet.title}'!{data_cells[0].coordinate}"
            for cell in data_cells[1:]:
                formula += f"-'{data1_sheet.title}'!{cell.coordinate}"

        if formula:
            cashflow.cell(row=row, column=week_col).value = formula
            print(f"✅ Mapped '{label}' to Cashflow col {week_col}: {formula}")


def map_d2w_values_to_cashflow(wb, d2_sheet, week, week_type="normal"):
    """Map D2W data to Cashflow sheet"""
    if "Cashflow" not in wb.sheetnames:
        print("Cashflow sheet not found")
        return

    cashflow = wb["Cashflow"]
    week_col = 3 + (week - 1)

    D2W_MAPPING = {
        "High Priority": (["Total Ads & miscellaneous services"], "B", "G"),
        "Zomato hyperpure Payment Adjustments": (["Total Hyperpure"], "B", "G"),
        "Service Fee Reversal": (["Service Fees Reversal"], "B", "G"),
        "Up-Time Pack Fee": (["Fees for Up-time pack"], "B", "G"),
        "Following Week Adjustments": (["Customer Compensation/Recoupment from previous week due to rejection", "Total adjustments from previous weeks"], "B", "G",)
    }

    print(f"\n🔍 D2W MAPPING - Scanning sheet '{d2_sheet.title}'...")

    for cashflow_label, (search_terms, search_col, value_col) in D2W_MAPPING.items():
        print(f"\n📊 Looking for '{cashflow_label}'...")

        found_row = None
        found_value = None

        for row in range(1, d2_sheet.max_row + 1):
            cell_value = str(d2_sheet[f"{search_col}{row}"].value or "").strip()

            for search_term in search_terms:
                if cell_value == search_term:
                    found_row = row
                    found_value = d2_sheet[f"{value_col}{row}"].value
                    print(f"  ✅ Found '{search_term}' at row {row}")
                    print(f"  📍 Value in column {value_col}{row}: {found_value}")
                    break

            if found_row:
                break

        if not found_value:
            print(f"  ❌ '{cashflow_label}' - No matching row found in column B")
            continue

        for cf_row in range(1, cashflow.max_row + 1):
            cf_label = str(cashflow.cell(row=cf_row, column=2).value or "").strip()

            if cf_label == cashflow_label:
                formula = f"='{d2_sheet.title}'!{value_col}{found_row}"
                cashflow.cell(row=cf_row, column=week_col).value = formula
                print(f"  ✅ Mapped to Cashflow row {cf_row}, col {week_col}: {formula}")
                break
        else:
            print(f"  ⚠️  Cashflow label '{cashflow_label}' not found in Cashflow sheet")


def map_commissionable_value_to_summary(summary_sheet, d1_sheet, week_num):
    """Map commissionable value to Summary sheet"""
    try:
        print(f"\n  🔍 Looking for Commissionable value column in D1W...")

        commissionable_col = None
        header_row = 5

        # ✅ Multiple possible header variations
        possible_headers = [
            "commissionable value (excludes customer gst)",
            "commissionable value (excludes customer gst)\n[(a) + (7) - (8) - (6) - (3)]",
            "commissionable value",
            "commissionable value\n(excludes customer gst)"
        ]

        for col_num in range(1, d1_sheet.max_column + 1):
            header = str(d1_sheet.cell(row=header_row, column=col_num).value or "").strip().lower()

            if "commissionable value" in header:
                commissionable_col = col_num
                print(f"  ✅ Found Commissionable value column: Col {col_num}")
                break

        if not commissionable_col:
            print(f"  ❌ Commissionable value column not found")
            return False

        commissionable_value = d1_sheet.cell(row=4, column=commissionable_col).value

        if commissionable_value is None:
            print(f"  ⚠️  No value in row 4, col {commissionable_col}")
            return False

        print(f"  📊 Commissionable value (row 4): {commissionable_value}")

        summary_col = 3 + (week_num - 1)

        summary_sheet.cell(row=19, column=summary_col).value = commissionable_value
        print(f"  ✅ Mapped to Summary[row=19, col={summary_col}]: {commissionable_value}")

        return True

    except Exception as e:
        print(f"  ❌ Error mapping commissionable value: {e}")
        import traceback
        traceback.print_exc()
        return False


def replace_month_in_sheets(wb, user_month):
    """Replace 'July' placeholder with actual month name"""
    sheets_to_update = ['Summary', 'Cashflow', 'Profit statement', 'Discrepancies']
    for sheet_name in sheets_to_update:
        if sheet_name not in wb.sheetnames:
            continue
        sheet = wb[sheet_name]
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and 'July' in cell.value:
                    cell.value = cell.value.replace('July', user_month)
    print(f"Replaced 'July' with '{user_month}' in all sheets")


def copy_logos_between_workbooks(template_path, target_wb, sheet_names_to_copy=None):
    """
    Copy logos/images from template to target workbook

    Args:
        template_path: Path to template Excel file
        target_wb: Target openpyxl workbook object (already loaded)
        sheet_names_to_copy: List of sheet names to copy logos from
    """
    if sheet_names_to_copy is None:
        sheet_names_to_copy = ["Summary"]

    try:
        from openpyxl import load_workbook

        print(f"\n📋 COPYING LOGOS FROM TEMPLATE...")
        print(f"   Template: {template_path}")

        # Load template (preserves all objects)
        template_wb = load_workbook(template_path)

        success_count = 0

        for sheet_name in sheet_names_to_copy:
            print(f"\n🔄 Sheet: '{sheet_name}'")

            if sheet_name not in template_wb.sheetnames:
                print(f"  ⚠️  Not found in template")
                continue

            if sheet_name not in target_wb.sheetnames:
                print(f"  ⚠️  Not found in target")
                continue

            template_sheet = template_wb[sheet_name]
            target_sheet = target_wb[sheet_name]

            # ✅ COPY THE DRAWING OBJECT (contains all images/logos)
            if hasattr(template_sheet, '_drawing') and template_sheet._drawing is not None:
                target_sheet._drawing = template_sheet._drawing
                print(f"  ✅ Logo copied!")
                success_count += 1
            else:
                print(f"  ℹ️  No drawings found")

        template_wb.close()

        if success_count > 0:
            print(f"\n✅ SUCCESS! ({success_count} sheet(s) updated with logos)")
            return True
        else:
            print(f"\n⚠️  No logos copied - check template")
            return False

    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()
        return False


def process_zomato_recon(
        invoice_folder_path,
        template_recon_path,
        output_path,
        client_name=None,
        month="October",
        first_week_start=None,  # ADD
        first_week_end=None,  # ADD
        last_week_start=None,  # ADD
        last_week_end=None,  # ADD
        bank_file_path=None
):
    """Zomato reconciliation engine"""
    try:
        folder = Path(invoice_folder_path)

        print(f"🔄 Copying template: {template_recon_path} → {output_path}")
        shutil.copy2(template_recon_path, output_path)
        print(f"✅ Template copied with logos!")

        recon = openpyxl.load_workbook(output_path)  # Load the COPY, not template
        print("Template loaded.")

        clear_all_D_sheets(recon)

        # NEW CODE:
        # ✅ Calculate week structure from user input
        print("\n📅 Calculating week structure...")
        week_structure = calculate_week_structure(
            month, first_week_start, first_week_end, last_week_start, last_week_end
        )

        print(f"\n✅ Week structure created: {len(week_structure)} weeks")
        for week in week_structure:
            spillover_info = ""
            if week['is_spillover_start']:
                spillover_info += " [Spillover Start]"
            if week['is_spillover_end']:
                spillover_info += " [Spillover End]"
            print(f"   Week {week['week_num']}: {week['label']}{spillover_info}")

        # ✅ Get all invoice files
        invoice_files = list(folder.glob("*.xlsx"))
        invoice_week_mapping = {}

        # ✅ Match each invoice to week structure
        print(f"\n📋 Found {len(invoice_files)} invoice files:")
        for fp in invoice_files:
            print(f"   - {fp.name}")

        print(f"\n🔍 Matching invoices to weeks...")

        for fp in invoice_files:
            filename = fp.name
            week_num = match_invoice_to_week(filename, week_structure, month)

            if week_num is None:
                print(f"⚠️  Skipping invoice (no week match): {filename}")
                continue

            invoice_week_mapping[fp] = week_num
            print(f"✅ Matched: {filename} → Week {week_num}")

        if not invoice_week_mapping:
            return {
                'success': False,
                'message': 'No invoices matched to week structure'
            }

        # ✅ Build week plan from matched invoices
        week_plan = []
        for week in week_structure:
            # Find invoice for this week
            invoice_fp = None
            for fp, mapped_week_num in invoice_week_mapping.items():
                if mapped_week_num == week['week_num']:
                    invoice_fp = fp
                    break

            if invoice_fp:
                week_plan.append({
                    'week_num': week['week_num'],
                    'invoice_fp': invoice_fp,
                    'week_label': week['label'],
                    'is_spillover_start': week['is_spillover_start'],
                    'is_spillover_end': week['is_spillover_end']
                })
                print(f"✅ Week {week['week_num']}: {week['label']} ← {invoice_fp.name}")
            else:
                # Week has no invoice - add placeholder
                week_plan.append({
                    'week_num': week['week_num'],
                    'invoice_fp': None,
                    'week_label': week['label'],
                    'is_spillover_start': week['is_spillover_start'],
                    'is_spillover_end': week['is_spillover_end']
                })
                print(f"⚠️  Week {week['week_num']}: {week['label']} - NO INVOICE")

        print(f"\n📋 Total weeks to process: {len(week_plan)}")
        for w in week_plan:
            print(f"  Week {w['week_num']}: {w['week_label']}")

        summary_sheet = ensure_sheet(recon, "Summary")
        if client_name:
            summary_sheet.cell(row=1, column=2).value = client_name

        for week_info in week_plan:
            col = 3 + (week_info['week_num'] - 1)
            summary_sheet.cell(row=4, column=col).value = week_info['week_label']
            print(f"✅ Week label '{week_info['week_label']}' → Summary[row=4, col={col}]")

        opening_spillover_value = 0
        closing_spillover_value = 0
        closing_week_num = 0

        for week_info in week_plan:
            fp = week_info['invoice_fp']
            week_num = week_info['week_num']

            # ✅ Skip if no invoice for this week
            if fp is None:
                print(f"\n--- Week {week_num}: {week_info['week_label']} - SKIPPED (no invoice) ---")
                continue

            print(f"\n--- Processing {fp.name} → Week {week_num} ---")

            wb_invoice = openpyxl.load_workbook(fp, data_only=True)

            try:
                possible_order_sheets = [
                    "Order Level", "Order level", "Order level Breakup",
                    "Order Details", "Order Summary", "Orders"
                ]
                ol_sheet = None
                for sheet_name in possible_order_sheets:
                    if sheet_name in wb_invoice.sheetnames:
                        ol_sheet = wb_invoice[sheet_name]
                        print(f"✅ ORDER SHEET: '{sheet_name}'")
                        break

                if ol_sheet:
                    d1 = ensure_sheet(recon, f"D1W{week_num}")

                    spillover_result = copy_data_with_spillover_filter(ol_sheet, d1, 7, month, week_info, None)

                    if spillover_result:
                        if spillover_result['opening_spillover'] != 0:
                            opening_spillover_value = spillover_result['opening_spillover']
                            print(f"  📝 Captured opening spillover: {opening_spillover_value}")

                        if spillover_result['closing_spillover'] != 0:
                            closing_spillover_value = spillover_result['closing_spillover']
                            closing_week_num = week_num
                            print(f"  📝 Captured closing spillover: {closing_spillover_value} (Week {week_num})")

                    perform_calculations_on_data1(recon, d1, week_num, output_path)
                    map_values_to_cashflow(recon, d1, week_num)

                    total_orders = count_total_orders_from_d1w(d1, header_row=5)
                    summary_col = 3 + (week_num - 1)
                    summary_sheet.cell(row=6, column=summary_col).value = total_orders
                    print(f"  ✅ Mapped Total Orders to Summary[row=6, col={summary_col}]: {total_orders}")

                    comp_orders = count_nonzero_compensation(d1, data_row=5)
                    summary_sheet.cell(row=12, column=summary_col).value = comp_orders
                    print(f"  ✅ Mapped Compensation Orders to Summary[row=12, col={summary_col}]: {comp_orders}")

                    map_commissionable_value_to_summary(summary_sheet, d1, week_num)

                else:
                    print(f"⚠️  No Order Level sheet found")

                # ✅ COPY D2W (Additions/Deductions)
                print(f"\n🔍 Looking for D2W sheet in {fp.name}...")
                print(f"  Available sheets: {wb_invoice.sheetnames}")

                # ✅ COPY D2W (Additions/Deductions)
                print(f"\n🔍 Looking for D2W sheet...")
                print(f"  Available sheets in invoice: {wb_invoice.sheetnames}")

                possible_d2_sheets = [
                    "Additions and Deductions",
                    "Addition Deductions Details",
                    "Additional and Deductions",
                    "Deductions"
                ]

                d2_source = None
                found_sheet_name = None
                for sheet_name in possible_d2_sheets:
                    if sheet_name in wb_invoice.sheetnames:
                        d2_source = wb_invoice[sheet_name]
                        found_sheet_name = sheet_name
                        print(f"✅ D2W SHEET FOUND: '{sheet_name}'")
                        break

                if d2_source:
                    d2 = ensure_sheet(recon, f"D2W{week_num}")
                    print(f"  📋 Target sheet: D2W{week_num}")

                    # Get dimensions
                    max_row = d2_source.max_row
                    max_col = d2_source.max_column
                    print(f"  📏 Source dimensions: {max_row} rows × {max_col} cols")

                    # Clear target sheet
                    if d2.max_row > 1:
                        d2.delete_rows(1, d2.max_row)
                        print(f"  🗑️  Cleared existing D2W{week_num} data")

                    # Copy ALL data from D2W source
                    copied_cells = 0
                    for r in range(1, max_row + 1):
                        for c in range(1, max_col + 1):
                            value = d2_source.cell(row=r, column=c).value
                            d2.cell(row=r, column=c).value = value
                            if value is not None:
                                copied_cells += 1

                    print(f"  ✅ Copied {copied_cells} cells to D2W{week_num}")

                    # ✅ VERIFY THE COPY
                    verify_row_1 = [d2.cell(row=1, column=c).value for c in range(1, min(5, max_col + 1))]
                    print(f"  🔍 Verification - First 4 cells of row 1: {verify_row_1}")

                    # Map D2W values to Cashflow
                    print(f"  🔗 Mapping D2W{week_num} to Cashflow...")
                    map_d2w_values_to_cashflow(recon, d2, week_num)

                else:
                    print(f"  ❌ No D2W sheet found!")
                    print(f"  💡 Available sheets were: {wb_invoice.sheetnames}")
                    print(f"  💡 Searched for: {possible_d2_sheets}")

            finally:
                wb_invoice.close()
                del wb_invoice
                gc.collect()
                print(f"  🧹 Closed invoice workbook")

        # ✅ SAVE WORKBOOK AFTER PROCESSING ALL WEEKS
        print(f"\n💾 Saving reconciliation workbook...")

        # Verify D2W sheets exist before saving
        d2w_sheets = [sh for sh in recon.sheetnames if sh.startswith("D2W")]
        print(f"  📋 D2W sheets in workbook: {d2w_sheets}")

        if d2w_sheets:
            # Check if they have data
            for d2w_name in d2w_sheets:
                d2w_sheet = recon[d2w_name]
                print(f"  📊 {d2w_name}: {d2w_sheet.max_row} rows, {d2w_sheet.max_column} cols")


        if opening_spillover_value != 0:
            print(f"\n📍 Mapping opening spillover: {opening_spillover_value}")
            cashflow = recon["Cashflow"]
            for row in range(1, cashflow.max_row + 1):
                label = str(cashflow.cell(row=row, column=2).value or "").strip()
                if label == "Opening Week Adjustments":
                    cashflow.cell(row=row, column=3).value = opening_spillover_value
                    print(f"  ✅ Mapped opening spillover to row {row}")
                    break

        if closing_spillover_value != 0 and closing_week_num > 0:
            print(f"\n📍 Mapping closing spillover: {closing_spillover_value}")
            cashflow = recon["Cashflow"]
            week_col = 3 + (closing_week_num - 1)
            for row in range(1, cashflow.max_row + 1):
                label = str(cashflow.cell(row=row, column=2).value or "").strip()
                if label == "Closing Week Adjustments":
                    cashflow.cell(row=row, column=week_col).value = closing_spillover_value
                    print(f"  ✅ Mapped closing spillover to row {row}, col {week_col}")
                    break

        replace_month_in_sheets(recon, month)

        recon.save(output_path)
        recon.close()
        print(f"\n✅ SUCCESS! Saved to: {output_path}")



        return {
            'success': True,
            'message': f'Processing complete! Generated reconciliation for {month}',
            'weeks_processed': len(week_plan)
        }

    except Exception as e:
        import traceback
        traceback.print_exc()
        return {
            'success': False,
            'message': f'Processing error: {str(e)}'
        }


if __name__ == "__main__":
    print("🚀 Zomato Reconciliation Tool\n")

    user_input = select_invoices_gui()

    if not user_input:
        print("❌ User cancelled or no files selected")
        exit()

    result = process_zomato_recon(
        invoice_folder_path=user_input['invoice_folder_path'],
        template_recon_path=user_input['template_recon_path'],
        output_path=user_input['output_path'],
        month="May"
    )

    if result['success']:
        print(f"\n✅ {result['message']}")
        print(f"📊 Weeks processed: {result['weeks_processed']}")
    else:
        print(f"\n❌ {result['message']}")
